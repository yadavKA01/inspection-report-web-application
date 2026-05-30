"""
Standalone YOLO detection API + static UI (does NOT use main.py).

Use this when another process on :9000 causes 404s or wrong `main` imports.

  cd AI_Engine
  python backend/serve_balloon.py

Repo layout: frontend/ (static UI) and backend/ (API + pipeline).

Then open the URL printed (default http://127.0.0.1:10000).

Returns JSON your frontend / .NET / Java can use to draw balloon circles:
  - detections[].bbox, class_name, confidence
  - drawing_annotations[] with id, BBox, TextPos (center), AnnotationType

Env:
  BALLOON_UI_PORT=10000   (optional)
  BALLOON_UI_HOST=127.0.0.1
"""
from __future__ import annotations

import asyncio
import base64
import hashlib
import hmac
import json
import math
import os
import re
import shutil
import sys
import threading
import uuid
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Optional

import cv2
import numpy as np
from pydantic import BaseModel, Field

_BACKEND_DIR = Path(__file__).resolve().parent
_APP_ROOT = _BACKEND_DIR.parent
_REPO_ROOT = _BACKEND_DIR.parent.parent

# Load .env from the backend directory (no-op if file or package is missing)
try:
    from dotenv import load_dotenv
    load_dotenv(_BACKEND_DIR / ".env")
except ImportError:
    pass
os.chdir(_BACKEND_DIR)
if str(_BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(_BACKEND_DIR))
for _p in ("Modules", "Dependencies", "Resources", ".Temp"):
    _d = str(_BACKEND_DIR / _p)
    if _d not in sys.path:
        sys.path.append(_d)

import config
import mongodb as db
from fastapi import Depends, FastAPI, File, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook

# ── Auth system (PostgreSQL / JWT / multi-tenant) ─────────────────────────
from auth.database import get_db, init_db
from auth.dependencies import (
    check_tenant_access,
    get_current_user,
    get_optional_user,
    require_active_subscription,
    require_balloon_write_access,
)
from auth.settings import auth_enabled, balloon_auth_disabled, database_configured, trial_days
from auth.models import Activity, Organization, RoleEnum, User
from auth.routes import router as auth_router
from auth.admin_routes import router as admin_router
from auth.session_routes import router as session_router
from sqlalchemy.exc import OperationalError, SQLAlchemyError
from sqlalchemy.orm import Session

config.InitConfiguration()
_Db = config.GetConfiguration("DATABASE")
if _Db:
    if _Db.get("URI"):
        db.Connect(uri=_Db["URI"])
    else:
        db.Connect(_Db.get("ADDRESS", "localhost"), _Db.get("PORT", 27017))
    if db.ping():
        print("[mongodb] Ping OK — database is reachable.")
    else:
        print(
            "[mongodb] WARNING: Ping failed. Check DATABASE.URI / MONGODB_URI, Atlas IP allowlist, "
            "and database user password (URL-encode special characters in the URI)."
        )
else:
    print(
        "[mongodb] WARNING: No DATABASE section in config — MongoDB not connected. "
        "Set DATABASE.URI (or ADDRESS/PORT) in config to enable database features."
    )

# Heavy deps (torch, ultralytics) — import lazily so Render binds $PORT before loading YOLO.
_tasks_mod = None


def _tasks():
    global _tasks_mod
    if _tasks_mod is None:
        from AutoBallooning import tasks as _tasks_mod
    return _tasks_mod


_UPLOAD_ROOT = _BACKEND_DIR / ".Temp" / "balloon_ui_uploads"
_UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)


def _resolve_ui_dir() -> Path:
    """
    Prefer frontend_padma/frontend (repo dev UI) when present so edits apply without copying
    to Resources. Set BALLOON_USE_BUNDLED_UI=1 to force Resources/balloon_ui.
    """
    fp = _REPO_ROOT / "frontend_padma" / "frontend"
    rs = _BACKEND_DIR / "Resources" / "balloon_ui"
    bundled = os.environ.get("BALLOON_USE_BUNDLED_UI", "").strip().lower() in ("1", "true", "yes")
    if bundled and (rs / "index.html").is_file():
        return rs
    if (fp / "index.html").is_file():
        return fp
    if (rs / "index.html").is_file():
        return rs
    return _APP_ROOT / "frontend"


_UI_DIR = _resolve_ui_dir()


def _cors_allow_origins() -> list[str]:
    """Explicit origins so credentialed cookies work; * is invalid with credentials."""
    raw = os.environ.get("BALLOON_CORS_ORIGINS", "").strip()
    if raw:
        return [x.strip() for x in raw.split(",") if x.strip()]
    default_ports = "3000,10000,9090," + os.environ.get("BALLOON_UI_PORT", "10000").strip()
    ports = sorted({p.strip() for p in default_ports.replace(" ", "").split(",") if p.strip()})
    out: list[str] = []
    for p in ports:
        for host in ("http://127.0.0.1", "http://localhost"):
            out.append(f"{host}:{p}")
    return out or ["http://127.0.0.1:10000"]


app = FastAPI(title="SmorX Balloon — detection API", version="1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_allow_origins(),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

_AUTH_DB_UNAVAILABLE = (
    "Authentication database is unavailable. For local development, comment out DATABASE_URL "
    "in backend/.env, or set AUTH_USE_SQLITE=1, then restart the server. "
    "For production, verify your Neon/PostgreSQL connection string."
)


@app.exception_handler(OperationalError)
async def _auth_operational_error_handler(request: Request, exc: OperationalError):
    print(f"[auth] DB error on {request.url.path}: {exc}")
    return JSONResponse(status_code=503, content={"detail": _AUTH_DB_UNAVAILABLE})


@app.exception_handler(SQLAlchemyError)
async def _auth_sqlalchemy_error_handler(request: Request, exc: SQLAlchemyError):
    if isinstance(exc, OperationalError):
        raise exc
    print(f"[auth] SQLAlchemy error on {request.url.path}: {exc}")
    return JSONResponse(status_code=503, content={"detail": _AUTH_DB_UNAVAILABLE})


# ── Register auth routers ──────────────────────────────────────────────────
app.include_router(auth_router)    # /auth/login, /auth/change-password, etc.
app.include_router(admin_router)   # /admin/organizations, /admin/engineers, etc.
app.include_router(session_router) # /activities/save, /activities, /activities/{id}


@app.on_event("startup")
def _startup():
    """Initialise PostgreSQL tables and seed super admin on first run."""
    if balloon_auth_disabled():
        print("[auth] SMORX_DISABLE_BALLOON_AUTH=1 — login not required (dev only).")
        return
    try:
        init_db()
        print("[auth] Auth database ready.")
    except Exception as exc:
        print(f"[auth] WARNING: DB init failed — {exc}")
        return
    if auth_enabled():
        print(f"[auth] Authentication ENABLED — {trial_days()}-day organization trial, then Razorpay payment.")
    if _deploy_safe_mode():
        print(
            "[detect] Render safe mode ON — YOLO only + limited crop OCR "
            f"(engine={_ocr_engine()}, max={_max_crop_ocr_count()}). "
            "Set BALLOON_RENDER_SAFE=0 for full Claude pipeline."
        )
        # NOTE: never load the YOLO model here — startup must finish fast so
        # uvicorn binds the port before Render's port-scan timeout. The model
        # is lazy-loaded on the first /api/v1/detect request instead.
        if os.environ.get("BALLOON_PRELOAD_YOLO", "").strip().lower() in ("1", "true", "yes", "on"):
            def _bg_preload():
                try:
                    _tasks().get_yolo_model()
                    print("[detect] YOLO model preloaded (background).")
                except Exception as exc:
                    print(f"[detect] YOLO preload skipped: {exc}")

            threading.Thread(target=_bg_preload, daemon=True).start()

_STATIC_DIR = _UI_DIR / "static"
if _STATIC_DIR.is_dir():
    app.mount("/assets", StaticFiles(directory=str(_STATIC_DIR)), name="balloon_assets")

print(f"[serve_balloon] UI: {_UI_DIR.resolve()}")
print("[serve_balloon] Open the app using ONE host only — e.g. http://127.0.0.1:9090")


class ExtractBalloonTextBody(BaseModel):
    """Client-sent crop (e.g. after drawing a manual box) for the same vision extract as auto-detect."""

    crop_jpeg_base64: str = Field(..., description="JPEG as data URL or raw base64")
    class_name: str = "Manual"


class PaymentVerifyRequest(BaseModel):
    razorpay_order_id: str
    razorpay_payment_id: str
    razorpay_signature: str


class InspectionReportRow(BaseModel):
    sno: int | str = ""
    balloon_number: str | int | float = ""
    reference_location: str = ""
    nominal: str = ""
    tol_low: str = ""
    tol_high: str = ""
    instrument: str = ""
    instrument_id: str = ""
    measured: list[str] = Field(default_factory=list)
    remarks: str = ""


class InspectionReportExportBody(BaseModel):
    part_number: str = ""
    part_name: str = ""
    revision: str = ""
    material: str = ""
    mass: str = ""
    finish: str = ""
    measured_col_count: int = 3
    rows: list[InspectionReportRow] = Field(default_factory=list)


def _fmt_tol_pdf(val: str) -> str:
    if val is None or str(val).strip() == "":
        return "—"
    return str(val).strip()


def _build_inspection_report_pdf(body: InspectionReportExportBody) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buff = BytesIO()
    page = landscape(A4)
    doc = SimpleDocTemplate(
        buff,
        pagesize=page,
        leftMargin=0.35 * inch,
        rightMargin=0.35 * inch,
        topMargin=0.4 * inch,
        bottomMargin=0.4 * inch,
    )
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle(
        "irCell",
        parent=styles["Normal"],
        fontSize=7,
        leading=8,
    )
    title_style = ParagraphStyle(
        "irTitle",
        parent=styles["Title"],
        fontSize=14,
        spaceAfter=6,
    )

    def cell(text: str) -> Paragraph:
        safe = (
            str(text if text is not None else "")
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )
        return Paragraph(safe or "—", cell_style)

    story: list = []
    story.append(Paragraph("Inspection Report", title_style))
    story.append(Spacer(1, 6))

    meta_table = Table(
        [
            [cell("Part Number"), cell(body.part_number), cell("Part Name"), cell(body.part_name)],
            [cell("Revision"), cell(body.revision), cell("Material"), cell(body.material)],
            [cell("Mass"), cell(body.mass), cell("Finish Treatment"), cell(body.finish)],
        ],
        colWidths=[1.05 * inch, 2.35 * inch, 1.05 * inch, 2.35 * inch],
    )
    meta_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#d6e8f7")),
                ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor("#9ec5e8")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#9ec5e8")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(meta_table)
    story.append(Spacer(1, 10))

    meas_n = max(1, int(body.measured_col_count or 1))
    headers = [
        "S.No",
        "Balloon No.",
        "Reference",
        "Nominal",
        "Tol (low)",
        "Tol (high)",
        "Instrument",
        "Instr. ID",
    ]
    for c in range(meas_n):
        headers.append(f"Measured {c + 1}")
    headers.append("Remarks")

    table_data = [[cell(h) for h in headers]]
    for row in body.rows:
        cells = [
            cell(row.sno),
            cell(row.balloon_number),
            cell(row.reference_location),
            cell(row.nominal or "—"),
            cell(_fmt_tol_pdf(row.tol_low)),
            cell(_fmt_tol_pdf(row.tol_high)),
            cell(row.instrument),
            cell(row.instrument_id),
        ]
        measured = row.measured or []
        for c in range(meas_n):
            cells.append(cell(measured[c] if c < len(measured) else ""))
        cells.append(cell(row.remarks))
        table_data.append(cells)

    usable_w = page[0] - doc.leftMargin - doc.rightMargin
    base_cols = 8
    extra = meas_n + 1
    w_sno, w_balloon, w_ref = 0.35, 0.55, 0.75
    w_nom, w_tol = 0.55, 0.5
    w_inst, w_inst_id = 0.65, 0.65
    w_meas = max(0.55, (usable_w / inch - w_sno - w_balloon - w_ref - w_nom - 2 * w_tol - w_inst - w_inst_id - 0.9) / meas_n)
    col_widths = [
        w_sno * inch,
        w_balloon * inch,
        w_ref * inch,
        w_nom * inch,
        w_tol * inch,
        w_tol * inch,
        w_inst * inch,
        w_inst_id * inch,
    ]
    col_widths.extend([w_meas * inch] * meas_n)
    col_widths.append(0.9 * inch)

    data_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    data_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f5a962")),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#d4e8f7")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#d4e8f7")]),
                ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor("#7eb8dc")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#7eb8dc")),
                ("LINEAFTER", (7, 0), (7, -1), 1.5, colors.HexColor("#e07a2f")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(data_table)
    doc.build(story)
    return buff.getvalue()


def _log_activity(
    db: Session,
    user: Optional[User],
    action_type: str,
    metadata: dict | None = None,
) -> None:
    """
    Persist an activity record to PostgreSQL.

    Engineers are stored under their own tenant_id.
    Super admin and unauthenticated (old UI) calls are silently skipped.
    Errors are swallowed so they never interrupt the main request.
    """
    if user is None or user.role == RoleEnum.super_admin or not user.tenant_id:
        return
    try:
        activity = Activity(
            tenant_id=user.tenant_id,
            user_id=user.id,
            action_type=action_type,
            action_metadata=metadata or {},
        )
        db.add(activity)
        db.commit()
    except Exception as exc:
        db.rollback()
        print(f"[auth] WARNING: Failed to log activity '{action_type}': {exc}")


def _html_no_cache(path: Path) -> FileResponse:
    resp = FileResponse(str(path), media_type="text/html; charset=utf-8")
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    return resp


def _bbox_tblr_key(det: dict) -> tuple[float, float]:
    """Sort key: top→bottom, then left→right using bbox top-left (y1, x1).

    Matches natural reading order on drawings better than center when boxes sit on one horizontal band.
    """
    bb = det.get("bbox")
    if not bb or len(bb) < 4:
        return (1e30, 1e30)
    x1, y1, x2, y2 = bb[0], bb[1], bb[2], bb[3]
    return (float(y1), float(x1))


def _reorder_detection_payload_tblr(payload: dict) -> None:
    """Align `detections` and `detections_full` to reading order (uses full-res boxes when both exist)."""
    dets = payload.get("detections") or []
    full = payload.get("detections_full")
    if not dets and not full:
        return
    if full is not None and len(full) == len(dets) and len(full) > 0:
        idx = sorted(range(len(full)), key=lambda i: _bbox_tblr_key(full[i]))
        payload["detections"] = [dets[i] for i in idx]
        payload["detections_full"] = [full[i] for i in idx]
    elif dets:
        payload["detections"] = sorted(dets, key=_bbox_tblr_key)
        if full is None:
            payload["detections_full"] = list(payload["detections"])


def _balloon_item_text(item: dict) -> str:
    parts = [
        item.get("nominal_value"),
        item.get("tolerance"),
        item.get("others"),
        item.get("detected_text"),
        item.get("multiplier_notation"),
        item.get("class_name"),
        item.get("raw_ocr"),
    ]
    return " ".join(str(p).strip() for p in parts if p is not None and str(p).strip())


def _parse_multiplier_count(text: str) -> int:
    if not text:
        return 0
    m = re.search(r"(\d+)\s*[xX×]", text)
    if not m:
        return 0
    n = int(m.group(1))
    return n if n >= 2 else 0


def _multiplier_count_from_item(item: dict) -> int:
    stored = item.get("multiplier_count")
    if isinstance(stored, int) and stored >= 2:
        return stored
    text_fields = " ".join(
        str(item.get(k) or "")
        for k in (
            "others",
            "nominal_value",
            "tolerance",
            "raw_ocr",
            "detected_text",
            "multiplier_notation",
        )
    )
    n = _parse_multiplier_count(text_fields)
    return n if n >= 2 else 0


def _drawing_label_for_ann(ann: dict) -> str:
    if not ann:
        return ""
    if ann.get("display_id") is not None and str(ann.get("display_id")).strip():
        d = str(ann["display_id"])
        m = re.match(r"^(\d+)\.\d+$", d)
        return m.group(1) if m else d
    if ann.get("parent_balloon_number") is not None:
        return str(ann["parent_balloon_number"])
    aid = str(ann.get("id", ""))
    m = re.match(r"^(\d+)\.\d+$", aid)
    return m.group(1) if m else aid


def _sync_balloon_items_from_detections(payload: dict) -> None:
    """Detected details / inspection report: one row per detection balloon number."""
    dets = list(payload.get("detections") or [])
    if not dets:
        return
    _ensure_drawing_annotations(payload)
    anns = list(payload.get("drawing_annotations") or [])
    items = list(payload.get("balloon_items") or [])
    by_di: dict[int, list] = {}

    for it in items:
        di = _detection_index_for_item(payload, it)
        if di is None:
            parent = _parent_balloon_number(it) or str(it.get("balloon_number") or "")
            for i, ann in enumerate(anns):
                if _drawing_label_for_ann(ann) == str(parent):
                    di = i
                    break
        if di is None:
            continue
        by_di.setdefault(int(di), []).append(it)

    out: list = []
    for i, d in enumerate(dets):
        group = by_di.get(i) or []
        if group:
            group.sort(
                key=lambda x: (
                    1 if _is_sub_balloon_item(x) else 0,
                    str(x.get("balloon_number") or ""),
                )
            )
            for it in group:
                row = dict(it)
                row["detection_index"] = i
                db = (dets[i] or {}).get("bbox") or []
                if len(db) >= 4:
                    row["bbox_pixels"] = list(db[:4])
                if not _is_sub_balloon_item(row):
                    lid = _drawing_label_for_ann(anns[i] if i < len(anns) else {})
                    if lid:
                        row["balloon_number"] = lid
                out.append(row)
            continue
        ann = anns[i] if i < len(anns) else {}
        bn = _drawing_label_for_ann(ann) or str(i + 1)
        bb = (d or {}).get("bbox") or []
        out.append(
            {
                "balloon_number": bn,
                "detection_index": i,
                "class_name": (d or {}).get("class_name") or "",
                "confidence": (d or {}).get("confidence", ""),
                "nominal_value": "",
                "tolerance": "",
                "others": "",
                "detected_text": "",
                "bbox_pixels": list(bb[:4]) if len(bb) >= 4 else [],
            }
        )
    payload["balloon_items"] = out


def _multiplier_notation(text: str, count: int) -> str:
    m = re.search(r"\b(\d+)\s*[xX]\b", text or "")
    if m:
        return re.sub(r"\s+", "", m.group(0))
    return f"{count}X" if count >= 2 else ""


def _sub_balloon_text_pos(bb: list, k: int, mult: int) -> list | None:
    """Offset balloon centers for nX features (e.g. 2X → .1 left, .2 right)."""
    if not bb or len(bb) < 4:
        return None
    x1, y1, x2, y2 = bb[0], bb[1], bb[2], bb[3]
    cx = (x1 + x2) / 2
    h = max(y2 - y1, 8.0)
    spread = max(28.0, (x2 - x1) * 0.4)
    offset_x = (k - (mult + 1) / 2) * spread
    if (x2 - x1) >= (y2 - y1) * 1.35:
        return [cx + offset_x, y2 + min(18.0, h * 0.5)]
    return [x2 + min(18.0, (x2 - x1) * 0.5), (y1 + y2) / 2]


def _bbox_for_balloon_row(payload: dict, i: int, it: dict) -> list | None:
    bp = it.get("bbox_pixels")
    if bp and len(bp) >= 4:
        return list(bp)
    anns = payload.get("drawing_annotations") or []
    dets = payload.get("detections") or []
    if i < len(anns):
        a = anns[i]
        bb = a.get("BBox") or a.get("bbox")
        if bb and len(bb) >= 4:
            return list(bb)
    if i < len(dets):
        bb = (dets[i] or {}).get("bbox")
        if bb and len(bb) >= 4:
            return list(bb)
    return None


def _is_sub_balloon_item(it: dict) -> bool:
    if it.get("is_sub_balloon"):
        return True
    bn = str(it.get("balloon_number") or "")
    return "." in bn and bn.split(".", 1)[1].isdigit()


def _parent_balloon_number(it: dict) -> str:
    p = it.get("parent_balloon_number")
    if p is not None and str(p).strip():
        return str(p)
    bn = str(it.get("balloon_number") or "")
    if "." in bn:
        return bn.split(".", 1)[0]
    return ""


def _build_canvas_balloon_annotations(payload: dict) -> list:
    """One canvas balloon per nX callout (label 15); table rows stay 15.1, 15.2."""
    items = list(payload.get("balloon_items") or [])
    anns = list(payload.get("drawing_annotations") or [])
    out: list = []
    drawn_parent: set[str] = set()

    for i, it in enumerate(items):
        bb = _bbox_for_balloon_row(payload, i, it)
        if not bb:
            continue
        if _is_sub_balloon_item(it):
            p = _parent_balloon_number(it)
            if not p or p in drawn_parent:
                continue
            drawn_parent.add(p)
            mult = int(it.get("multiplier_count") or 2)
            base = dict(anns[i]) if i < len(anns) else {}
            tp = _sub_balloon_text_pos(bb, 1, mult)
            row = {
                **base,
                "id": p,
                "display_id": p,
                "BBox": bb,
                "_sub_balloon": True,
                "parent_balloon_number": p,
                "canvas_skip": False,
            }
            if tp:
                row["TextPos"] = tp
            out.append(row)
            continue
        bn = str(it.get("balloon_number", i + 1))
        base = dict(anns[i]) if i < len(anns) else {}
        out.append(
            {
                **base,
                "id": bn,
                "display_id": bn,
                "BBox": bb,
                "_sub_balloon": False,
                "canvas_skip": False,
            }
        )
    return out


def _sync_sub_balloon_canvas_meta(payload: dict) -> None:
    """On drawing show whole balloon # (15); table rows stay 15.1, 15.2."""
    items = list(payload.get("balloon_items") or [])
    anns = list(payload.get("drawing_annotations") or [])
    groups: dict[str, list[int]] = {}
    for i, it in enumerate(items):
        if not it.get("is_sub_balloon"):
            continue
        p = str(it.get("parent_balloon_number") or "")
        if not p:
            continue
        groups.setdefault(p, []).append(i)
    for p, idxs in groups.items():
        mult = len(idxs)
        for j, idx in enumerate(idxs):
            if idx >= len(anns):
                continue
            ann = dict(anns[idx])
            ann["display_id"] = p
            ann.setdefault("parent_balloon_number", p)
            ann["canvas_skip"] = j > 0
            if not ann["canvas_skip"]:
                tp = _sub_balloon_text_pos(ann.get("BBox"), 1, mult)
                if tp:
                    ann["TextPos"] = tp
            anns[idx] = ann
    payload["drawing_annotations"] = anns


def _prune_parent_balloons_with_subs(payload: dict) -> None:
    """Table only: drop parent row when subs exist (15.1, 15.2). Drawing unchanged."""
    items = list(payload.get("balloon_items") or [])
    if not items:
        return

    bases_with_subs: set[str] = set()
    for it in items:
        if not _is_sub_balloon_item(it):
            continue
        p = _parent_balloon_number(it)
        if p:
            bases_with_subs.add(p)

    if not bases_with_subs:
        return

    payload["balloon_items"] = [
        it
        for it in items
        if _is_sub_balloon_item(it)
        or (
            not it.get("is_parent_balloon")
            and str(it.get("balloon_number") or "") not in bases_with_subs
            and _multiplier_count_from_item(it) < 2
        )
    ]


def _detection_index_for_item(payload: dict, it: dict) -> int | None:
    di = it.get("detection_index")
    if di is not None:
        return int(di)
    bp = it.get("bbox_pixels")
    if not bp or len(bp) < 4:
        return None
    for i, d in enumerate(payload.get("detections") or []):
        bb = (d or {}).get("bbox")
        if bb and len(bb) >= 4 and list(bb[:4]) == list(bp[:4]):
            return i
    return None


def _bbox_near_duplicate(primary_bb: list, this_bb: list) -> bool:
    if not primary_bb or not this_bb or len(primary_bb) < 4 or len(this_bb) < 4:
        return False
    if _bbox_overlap(primary_bb, this_bb) >= 0.15:
        return True
    y_overlap = min(primary_bb[3], this_bb[3]) - max(primary_bb[1], this_bb[1])
    span_y = max(
        primary_bb[3] - primary_bb[1],
        this_bb[3] - this_bb[1],
        12.0,
    )
    if y_overlap < span_y * 0.35:
        return False
    span = max(
        primary_bb[2] - primary_bb[0],
        primary_bb[3] - primary_bb[1],
        this_bb[2] - this_bb[0],
        this_bb[3] - this_bb[1],
        48.0,
    )
    pcx = (primary_bb[0] + primary_bb[2]) / 2
    pcy = (primary_bb[1] + primary_bb[3]) / 2
    tcx = (this_bb[0] + this_bb[2]) / 2
    tcy = (this_bb[1] + this_bb[3]) / 2
    return math.hypot(pcx - tcx, pcy - tcy) < span * 2.5


def _bbox_near_multiplier_duplicate(primary_bb: list, this_bb: list) -> bool:
    """Only hide duplicate YOLO slot on the same nX callout (high overlap), not neighboring dimensions."""
    if not primary_bb or not this_bb or len(primary_bb) < 4 or len(this_bb) < 4:
        return False
    return _bbox_overlap(primary_bb, this_bb) >= float(
        os.environ.get("BALLOON_MULTIPLIER_CANVAS_SKIP_IOU", "0.72")
    )


def _multiplier_primary_detection_indices(payload: dict) -> set[int]:
    primaries: set[int] = set()
    for it in payload.get("balloon_items") or []:
        if not _is_sub_balloon_item(it):
            continue
        di = _detection_index_for_item(payload, it)
        if di is not None:
            primaries.add(int(di))
    for i, _d in enumerate(payload.get("detections") or []):
        row = None
        for it in payload.get("balloon_items") or []:
            if _detection_index_for_item(payload, it) == i:
                row = it
                break
        if row is None:
            items = payload.get("balloon_items") or []
            row = items[i] if i < len(items) else {}
        if _multiplier_count_from_item(row or {}) >= 2:
            primaries.add(int(i))
    return primaries


def _remove_duplicate_yolo_detections_near_multiplier(payload: dict) -> None:
    """Remove extra YOLO slots beside nX callouts (no balloon 16/17 on drawing)."""
    dets = list(payload.get("detections") or [])
    if len(dets) < 2:
        return
    _ensure_drawing_annotations(payload)
    primaries = _multiplier_primary_detection_indices(payload)
    if not primaries:
        return

    to_remove: set[int] = set()
    for primary_di in primaries:
        primary_bb = _bbox_for_balloon_row(
            payload,
            primary_di,
            next(
                (
                    it
                    for it in (payload.get("balloon_items") or [])
                    if _detection_index_for_item(payload, it) == primary_di
                ),
                {},
            ),
        )
        if not primary_bb:
            continue
        for j in range(len(dets)):
            if j == primary_di or j in primaries or j in to_remove:
                continue
            this_bb = _bbox_for_balloon_row(payload, j, {})
            if this_bb and _bbox_near_multiplier_duplicate(primary_bb, this_bb):
                to_remove.add(j)

    if not to_remove:
        return

    map_old_to_new: dict[int, int] = {}
    new_dets: list = []
    new_anns: list = []
    anns = list(payload.get("drawing_annotations") or [])
    for i, d in enumerate(dets):
        if i in to_remove:
            continue
        map_old_to_new[i] = len(new_dets)
        new_dets.append(d)
        if i < len(anns):
            new_anns.append(anns[i])

    payload["detections"] = new_dets
    payload["drawing_annotations"] = new_anns
    payload["count"] = len(new_dets)
    payload.pop("canvas_balloon_annotations", None)

    new_items: list = []
    for it in payload.get("balloon_items") or []:
        di = _detection_index_for_item(payload, it)
        if di is not None and di in to_remove:
            continue
        row = dict(it)
        if di is not None and di in map_old_to_new:
            row["detection_index"] = map_old_to_new[di]
        new_items.append(row)
    payload["balloon_items"] = new_items


def _bbox_overlap(a: list, b: list) -> float:
    if not a or not b or len(a) < 4 or len(b) < 4:
        return 0.0
    x1, y1 = max(a[0], b[0]), max(a[1], b[1])
    x2, y2 = min(a[2], b[2]), min(a[3], b[3])
    if x2 <= x1 or y2 <= y1:
        return 0.0
    inter = (x2 - x1) * (y2 - y1)
    area_a = max(1.0, (a[2] - a[0]) * (a[3] - a[1]))
    area_b = max(1.0, (b[2] - b[0]) * (b[3] - b[1]))
    return inter / min(area_a, area_b)


_DETECTION_CLASS_MIN_CONF: dict[str, float] = {
    "Dimensions": 0.08,
    "GDnT": 0.07,
    "Notes": 0.10,
    "Surface_Finish_Symbols": 0.09,
    "Special_Characteristics": 0.09,
}
_DEFAULT_MIN_CONF = 0.08


def _bbox_ink_ratio(gray: np.ndarray, x1: int, y1: int, x2: int, y2: int) -> float:
    """Fraction of dark pixels inside bbox (dimension text / lines)."""
    h, w = gray.shape[:2]
    x1 = max(0, min(x1, w - 1))
    x2 = max(x1 + 1, min(x2, w))
    y1 = max(0, min(y1, h - 1))
    y2 = max(y1 + 1, min(y2, h))
    roi = gray[y1:y2, x1:x2]
    if roi.size == 0:
        return 0.0
    return float(np.count_nonzero(roi < 175)) / float(roi.size)


def _strict_detection_filter_enabled() -> bool:
    return os.environ.get("BALLOON_STRICT_DETECTION_FILTER", "").strip().lower() in (
        "1",
        "true",
        "yes",
    )


def _is_spurious_detection_light(
    d: dict,
    img_w: int,
    img_h: int,
    gray: np.ndarray | None = None,
) -> bool:
    """Light filter: drop only obvious bad boxes (keeps most YOLO output)."""
    bb = (d or {}).get("bbox") or []
    if len(bb) < 4:
        return True
    x1, y1, x2, y2 = int(bb[0]), int(bb[1]), int(bb[2]), int(bb[3])
    if x2 <= x1 or y2 <= y1:
        return True
    bw, bh = x2 - x1, y2 - y1
    conf = float((d or {}).get("confidence") or 0.0)
    if conf < 0.05:
        return True
    if bw < 6 or bh < 6:
        return True
    area_frac = (bw * bh) / max(1, img_w * img_h)
    if area_frac > 0.06:
        return True
    cls = str((d or {}).get("class_name") or "Dimensions").strip()
    if cls == "Dimensions" and area_frac > 0.035:
        return True
    if gray is not None and area_frac > 0.02:
        ink = _bbox_ink_ratio(gray, x1, y1, x2, y2)
        if ink < 0.008:
            return True
    return False


def _is_spurious_detection(
    d: dict,
    img_w: int,
    img_h: int,
    gray: np.ndarray | None = None,
) -> bool:
    """Strict filter: border/title-block/empty-area (set BALLOON_STRICT_DETECTION_FILTER=1)."""
    bb = (d or {}).get("bbox") or []
    if len(bb) < 4:
        return True
    x1, y1, x2, y2 = int(bb[0]), int(bb[1]), int(bb[2]), int(bb[3])
    if x2 <= x1 or y2 <= y1:
        return True

    bw, bh = x2 - x1, y2 - y1
    cx, cy = (x1 + x2) / 2.0, (y1 + y2) / 2.0
    cls = str((d or {}).get("class_name") or "Dimensions").strip()
    conf = float((d or {}).get("confidence") or 0.0)
    min_conf = _DETECTION_CLASS_MIN_CONF.get(cls, _DEFAULT_MIN_CONF)
    if conf < min_conf:
        return True

    img_area = max(1, img_w * img_h)
    area_frac = (bw * bh) / img_area

    if bw < 10 or bh < 10:
        return True
    if area_frac > 0.035:
        return True
    if cls == "Dimensions" and area_frac > 0.022:
        return True
    if cls == "Dimensions" and (bw > img_w * 0.28 or bh > img_h * 0.12):
        return True

    if gray is not None and conf < 0.35:
        ink = _bbox_ink_ratio(gray, x1, y1, x2, y2)
        margin = 0.025
        on_border = (
            cx < img_w * margin
            or cx > img_w * (1.0 - margin)
            or cy < img_h * margin
            or cy > img_h * (1.0 - margin)
        )
        if on_border and ink < 0.02 and area_frac > 0.008:
            return True
        if cls in ("Dimensions", "GDnT") and ink < 0.018:
            return True
        if cls == "Notes" and ink < 0.022 and area_frac > 0.006:
            return True

    return False


_CLASS_KEEP_RANK = {
    "Dimensions": 0,
    "GDnT": 1,
    "Notes": 2,
    "Datums": 3,
    "Surface_Finish_Symbols": 4,
    "Welding_Symbols": 5,
    "Special_Characteristics": 6,
    "Revision_Table": 7,
    "Title_Block": 8,
    "Miscellaneous": 9,
}


def _is_vision_sourced(d: dict) -> bool:
    return "vision" in str((d or {}).get("source") or "").lower()


def _detection_keep_rank(d: dict) -> tuple:
    """Lower = prefer keeping this box when two overlap (YOLO beats Claude)."""
    cls = str((d or {}).get("class_name") or "Miscellaneous").strip()
    return (
        1 if _is_vision_sourced(d) else 0,
        _CLASS_KEEP_RANK.get(cls, 5),
        -float((d or {}).get("confidence") or 0.0),
    )


def _is_spurious_detection_medium(
    d: dict,
    img_w: int,
    img_h: int,
    gray: np.ndarray | None = None,
) -> bool:
    """Drop empty lines, extension lines, and blank-area vision boxes after YOLO+Claude merge."""
    if _is_spurious_detection_light(d, img_w, img_h, gray):
        return True
    bb = (d or {}).get("bbox") or []
    if len(bb) < 4:
        return True
    x1, y1, x2, y2 = int(bb[0]), int(bb[1]), int(bb[2]), int(bb[3])
    bw, bh = x2 - x1, y2 - y1
    cls = str((d or {}).get("class_name") or "Dimensions").strip()
    conf = float((d or {}).get("confidence") or 0.0)
    area_frac = (bw * bh) / max(1, img_w * img_h)
    aspect = max(bw, bh) / max(1, min(bw, bh))

    if gray is not None:
        ink = _bbox_ink_ratio(gray, x1, y1, x2, y2)
        if cls in ("Dimensions", "GDnT") and ink < 0.014:
            return True
        if cls == "Notes" and ink < 0.028:
            return True
        if cls == "Miscellaneous" and ink < 0.032:
            return True
        if _is_vision_sourced(d) and conf < 0.65 and ink < 0.022:
            return True
        # Thin boxes on lines (extension lines, grid) not on text
        if aspect >= 5.5 and min(bw, bh) < max(14, int(img_h * 0.012)) and ink < 0.02:
            return True
        if cls in ("Dimensions", "Miscellaneous") and aspect >= 4.0 and ink < 0.018:
            return True

    if cls == "Miscellaneous" and area_frac < 0.00015 and conf < 0.7:
        return True
    if _is_vision_sourced(d) and cls == "Miscellaneous" and area_frac > 0.004 and conf < 0.6:
        return True
    return False


def _boxes_are_distinct_dimensions(d_i: dict, d_j: dict) -> bool:
    """Keep separate vertical/horizontal dims on the same line (different values)."""
    if str((d_i or {}).get("class_name") or "") != "Dimensions":
        return False
    if str((d_j or {}).get("class_name") or "") != "Dimensions":
        return False
    bb_i = (d_i or {}).get("bbox") or []
    bb_j = (d_j or {}).get("bbox") or []
    if len(bb_i) < 4 or len(bb_j) < 4:
        return False
    wi, hi = bb_i[2] - bb_i[0], bb_i[3] - bb_i[1]
    wj, hj = bb_j[2] - bb_j[0], bb_j[3] - bb_j[1]
    cxi, cyi = _bbox_center(bb_i)
    cxj, cyj = _bbox_center(bb_j)
    ori_i = (d_i or {}).get("dimension_orientation") or _dimension_callout_orientation(wi, hi)
    ori_j = (d_j or {}).get("dimension_orientation") or _dimension_callout_orientation(wj, hj)
    if ori_i == "vertical" and ori_j == "vertical":
        if abs(cxi - cxj) < max(wi, wj) * 0.55 and abs(cyi - cyj) > min(hi, hj) * 0.32:
            return True
    if ori_i == "horizontal" and ori_j == "horizontal":
        if abs(cyi - cyj) < max(hi, hj) * 0.55 and abs(cxi - cxj) > min(wi, wj) * 0.32:
            return True
    return False


def _dedupe_detections_by_overlap(dets: list, iou_keep: float = 0.45) -> list:
    """Drop overlapping boxes; keep YOLO / Dimensions over Claude / Miscellaneous."""
    if len(dets) < 2:
        return dets
    ranked = sorted(range(len(dets)), key=lambda i: _detection_keep_rank(dets[i]))
    kept: list[int] = []
    for i in ranked:
        bb_i = (dets[i] or {}).get("bbox") or []
        if len(bb_i) < 4:
            continue
        dup = False
        for j in kept:
            bb_j = (dets[j] or {}).get("bbox") or []
            if len(bb_j) < 4:
                continue
            if _bbox_overlap(bb_i, bb_j) >= iou_keep:
                if _boxes_are_distinct_dimensions(dets[i], dets[j]):
                    continue
                dup = True
                break
        if not dup:
            kept.append(i)
    kept.sort()
    return [dets[i] for i in kept]


def _merge_notes_blocks(dets: list) -> list:
    """Merge multiple Notes boxes in the same notes column (fewer balloons on empty lines)."""
    notes_idx = [i for i, d in enumerate(dets) if str(d.get("class_name") or "") == "Notes"]
    if len(notes_idx) < 2:
        return dets
    out = [dict(d) for d in dets]
    used: set[int] = set()

    def h_overlap(a: list, b: list) -> bool:
        left = max(a[0], b[0])
        right = min(a[2], b[2])
        w = min(a[2] - a[0], b[2] - b[0])
        return w > 0 and (right - left) / w >= 0.35

    for i in notes_idx:
        if i in used:
            continue
        group = [i]
        bb_i = out[i]["bbox"]
        for j in notes_idx:
            if j <= i or j in used:
                continue
            bb_j = out[j]["bbox"]
            if len(bb_j) < 4 or len(bb_i) < 4:
                continue
            cy_i = (bb_i[1] + bb_i[3]) / 2
            cy_j = (bb_j[1] + bb_j[3]) / 2
            vgap = abs(cy_i - cy_j)
            line_h = max(bb_i[3] - bb_i[1], bb_j[3] - bb_j[1], 8)
            if h_overlap(bb_i, bb_j) and vgap < line_h * 3.5:
                group.append(j)
        if len(group) < 2:
            continue
        xs, ys = [], []
        confs = []
        for gi in group:
            b = out[gi].get("bbox") or []
            if len(b) < 4:
                continue
            xs.extend([b[0], b[2]])
            ys.extend([b[1], b[3]])
            confs.append(float(out[gi].get("confidence") or 0))
            used.add(gi)
        if not xs or not ys:
            continue
        anchor = group[0]
        out[anchor]["bbox"] = [min(xs), min(ys), max(xs), max(ys)]
        out[anchor]["confidence"] = max(confs)
        for gi in group[1:]:
            out[gi] = None
    merged = [d for d in out if d is not None]
    return merged


def _bbox_center(bb: list) -> tuple[float, float]:
    return ((bb[0] + bb[2]) / 2.0, (bb[1] + bb[3]) / 2.0)


def _merge_class_boxes(dets: list, class_name: str, max_gap_px: float) -> list:
    """Merge same-class boxes that are close (e.g. many Title_Block field boxes → one)."""
    idxs = [i for i, d in enumerate(dets) if str(d.get("class_name") or "") == class_name]
    if len(idxs) < 2:
        return dets
    out = [dict(d) for d in dets]
    used: set[int] = set()
    for i in idxs:
        if i in used:
            continue
        group = [i]
        bb_i = out[i].get("bbox") or []
        if len(bb_i) < 4:
            continue
        cxi, cyi = _bbox_center(bb_i)
        for j in idxs:
            if j <= i or j in used:
                continue
            bb_j = out[j].get("bbox") or []
            if len(bb_j) < 4:
                continue
            cxj, cyj = _bbox_center(bb_j)
            if math.hypot(cxi - cxj, cyi - cyj) <= max_gap_px:
                group.append(j)
        if len(group) < 2:
            continue
        xs, ys, confs = [], [], []
        for gi in group:
            b = out[gi].get("bbox") or []
            if len(b) < 4:
                continue
            xs.extend([b[0], b[2]])
            ys.extend([b[1], b[3]])
            confs.append(float(out[gi].get("confidence") or 0))
            used.add(gi)
        if not xs or not ys:
            continue
        anchor = group[0]
        out[anchor]["bbox"] = [min(xs), min(ys), max(xs), max(ys)]
        out[anchor]["confidence"] = max(confs)
        for gi in group[1:]:
            out[gi] = None
    return [d for d in out if d is not None]


def _drop_miscellaneous_title_corner(dets: list, img_w: int, img_h: int) -> list:
    """Remove tiny Miscellaneous boxes in the title-block corner (false vision hits)."""
    if img_w < 1 or img_h < 1:
        return dets
    kept = []
    for d in dets:
        cls = str((d or {}).get("class_name") or "")
        if cls != "Miscellaneous":
            kept.append(d)
            continue
        bb = (d or {}).get("bbox") or []
        if len(bb) < 4:
            continue
        cx, cy = _bbox_center(bb)
        area_frac = ((bb[2] - bb[0]) * (bb[3] - bb[1])) / max(1, img_w * img_h)
        in_corner = cx > img_w * 0.52 and cy > img_h * 0.58
        if in_corner and area_frac < 0.012:
            continue
        kept.append(d)
    return kept


def _post_detect_cleanup_enabled() -> bool:
    return os.environ.get("BALLOON_POST_DETECT_CLEANUP", "1").strip().lower() not in (
        "0",
        "false",
        "no",
        "off",
    )


def _cleanup_detection_payload(payload: dict) -> None:
    """
    After YOLO + Claude gap-fill: keep ALL YOLO detections (balloons already assigned).
    Only filter/dedupe Claude additions — never collapse 28 YOLO boxes down to a few.
    """
    if not _post_detect_cleanup_enabled():
        return
    full = list(payload.get("detections_full") or payload.get("detections") or [])
    if not full:
        return
    before = len(full)
    infer_path = payload.get("infer_image_path")
    gray = None
    w_full, h_full = 0, 0
    if infer_path and Path(str(infer_path)).is_file():
        gray = cv2.imread(str(infer_path), cv2.IMREAD_GRAYSCALE)
        if gray is not None:
            h_full, w_full = gray.shape[:2]
    if w_full < 1:
        w_full = int(payload.get("width") or 0)
    if h_full < 1:
        h_full = int(payload.get("height") or 0)

    yolo_dets = [d for d in full if not _is_vision_sourced(d)]
    vision_dets = [d for d in full if _is_vision_sourced(d)]
    # Never dedupe YOLO — every YOLO detection keeps its balloon (user requirement).

    vision_dets = [
        d for d in vision_dets if not _is_spurious_detection_medium(d, w_full, h_full, gray)
    ]
    vision_dets = _drop_miscellaneous_title_corner(vision_dets, w_full, h_full)

    merged = list(yolo_dets)
    v_merge = float(os.environ.get("BALLOON_VISION_MERGE_IOU", "0.40"))
    for vd in vision_dets:
        bb = vd.get("bbox") or []
        if len(bb) < 4:
            continue
        if _max_bbox_iou_with_list(bb, merged) >= v_merge:
            continue
        merged.append(vd)

    merged = _merge_notes_blocks(merged)
    gap_tb = max(120.0, w_full * 0.08) if w_full > 0 else 120.0
    gap_rev = max(80.0, w_full * 0.05) if w_full > 0 else 80.0
    merged = _merge_class_boxes(merged, "Title_Block", gap_tb)
    merged = _merge_class_boxes(merged, "Revision_Table", gap_rev)

    _apply_detections_full_to_payload(payload, merged)
    payload["yolo_preserved_count"] = len(yolo_dets)
    payload["detections_cleaned"] = max(0, before - len(merged))
    payload["count"] = len(payload.get("detections") or [])


def _filter_detection_payload(payload: dict) -> None:
    """Remove false-positive detections (light by default; strict via env)."""
    full = list(payload.get("detections_full") or [])
    disp = list(payload.get("detections") or [])
    if not full:
        return

    strict = _strict_detection_filter_enabled()
    reject_fn = _is_spurious_detection if strict else _is_spurious_detection_light
    dedupe_iou = 0.45 if strict else 0.90

    infer_path = payload.get("infer_image_path")
    gray = None
    w_full, h_full = 0, 0
    if infer_path and Path(str(infer_path)).is_file():
        gray = cv2.imread(str(infer_path), cv2.IMREAD_GRAYSCALE)
        if gray is not None:
            h_full, w_full = gray.shape[:2]

    if w_full < 1 or h_full < 1:
        w_full = int(payload.get("width") or 0)
        h_full = int(payload.get("height") or 0)

    kept_full: list = []
    kept_disp: list = []
    removed = 0
    for i, d in enumerate(full):
        if reject_fn(d, w_full, h_full, gray):
            removed += 1
            continue
        kept_full.append(d)
        if i < len(disp):
            kept_disp.append(disp[i])

    kept_full = _dedupe_detections_by_overlap(kept_full)
    sc = float(payload.get("width") or w_full) / float(w_full) if w_full > 0 else 1.0
    kept_disp = []
    for d in kept_full:
        bb = d.get("bbox") or []
        if len(bb) < 4:
            continue
        if abs(sc - 1.0) < 0.001:
            kept_disp.append(d)
        else:
            kept_disp.append(
                {
                    "class_name": d.get("class_name"),
                    "confidence": d.get("confidence"),
                    "bbox": [
                        int(bb[0] * sc),
                        int(bb[1] * sc),
                        int(bb[2] * sc),
                        int(bb[3] * sc),
                    ],
                }
            )

    payload["detections_full"] = kept_full
    payload["detections"] = kept_disp or kept_full
    payload["count"] = len(payload["detections"])
    payload["detections_removed"] = removed
    payload["detection_filter_mode"] = "strict" if strict else "light"


def _bbox_ink_centroid_y(gray: np.ndarray, x1: int, y1: int, x2: int, y2: int) -> float | None:
    h, w = gray.shape[:2]
    x1 = max(0, min(x1, w - 1))
    x2 = max(x1 + 1, min(x2, w))
    y1 = max(0, min(y1, h - 1))
    y2 = max(y1 + 1, min(y2, h))
    roi = gray[y1:y2, x1:x2]
    if roi.size == 0:
        return None
    mask = roi < 175
    if not np.any(mask):
        return None
    rows = np.where(mask)[0] + y1
    return float(np.mean(rows))


def _bbox_ink_centroid_x(gray: np.ndarray, x1: int, y1: int, x2: int, y2: int) -> float | None:
    """Horizontal center of mass of dark pixels in bbox (for balloon side)."""
    h, w = gray.shape[:2]
    x1 = max(0, min(x1, w - 1))
    x2 = max(x1 + 1, min(x2, w))
    y1 = max(0, min(y1, h - 1))
    y2 = max(y1 + 1, min(y2, h))
    roi = gray[y1:y2, x1:x2]
    if roi.size == 0:
        return None
    mask = roi < 175
    if not np.any(mask):
        return None
    cols = np.where(mask)[1] + x1
    return float(np.mean(cols))


def _dimension_callout_orientation(w: float, h: float) -> str:
    if h >= w * 1.15:
        return "vertical"
    if w >= h * 1.15:
        return "horizontal"
    return "square"


def _refine_dimension_callout_bbox(
    d: dict, gray: np.ndarray | None, img_w: int, img_h: int
) -> dict:
    """
    Expand Dimension boxes so vertical length callouts (line + arrows + rotated text)
    are fully enclosed — improves balloon targeting.
    """
    bb = list((d or {}).get("bbox") or [])
    if len(bb) < 4:
        return d
    x1, y1, x2, y2 = int(bb[0]), int(bb[1]), int(bb[2]), int(bb[3])
    if x2 <= x1 or y2 <= y1:
        return d
    w, h = x2 - x1, y2 - y1
    cls = str((d or {}).get("class_name") or "").strip()
    if cls != "Dimensions":
        return d

    ori = _dimension_callout_orientation(float(w), float(h))
    if ori == "vertical":
        pad_left = max(22, int(h * 0.38))
        pad_right = max(14, int(h * 0.18))
        pad_y = max(12, int(h * 0.12))
        x1 = max(0, x1 - pad_left)
        x2 = min(img_w, x2 + pad_right)
        y1 = max(0, y1 - pad_y)
        y2 = min(img_h, y2 + pad_y)
    elif ori == "horizontal":
        pad_x = max(8, int(w * 0.08))
        pad_top = max(16, int(w * 0.28))
        pad_bottom = max(10, int(w * 0.14))
        x1 = max(0, x1 - pad_x)
        x2 = min(img_w, x2 + pad_x)
        y1 = max(0, y1 - pad_top)
        y2 = min(img_h, y2 + pad_bottom)
    else:
        pad = max(10, int(min(w, h) * 0.12))
        x1 = max(0, x1 - pad)
        y1 = max(0, y1 - pad)
        x2 = min(img_w, x2 + pad)
        y2 = min(img_h, y2 + pad)

    out = dict(d)
    out["bbox"] = [x1, y1, x2, y2]
    out["dimension_orientation"] = ori
    if gray is not None:
        cx_ink = _bbox_ink_centroid_x(gray, x1, y1, x2, y2)
        if cx_ink is not None:
            box_cx = (x1 + x2) / 2.0
            if ori == "vertical":
                out["balloon_side"] = "right" if cx_ink < box_cx else "left"
            elif ori == "horizontal":
                cy_ink = _bbox_ink_centroid_y(gray, x1, y1, x2, y2)
                if cy_ink is not None:
                    out["balloon_side"] = "below" if cy_ink < (y1 + y2) / 2.0 else "above"
    return out


def _dimension_bbox_refine_enabled() -> bool:
    """Expand bboxes for vertical/horizontal dim lines (on by default)."""
    raw = os.environ.get("BALLOON_DIMENSION_BBOX_REFINE", "1").strip().lower()
    return raw not in ("0", "false", "no", "off")


def _balloon_placement_mode() -> str:
    """legacy = AutoBallooning getBalloonCoordinates (default); tight = beside bbox."""
    return (os.environ.get("BALLOON_PLACEMENT", "legacy") or "legacy").strip().lower()


def _apply_legacy_balloon_coordinates(payload: dict, image_path: str) -> None:
    """Use the same balloon placement as full AutoBallooning (drawing3-style)."""
    anns = list(payload.get("drawing_annotations") or [])
    if not anns or not image_path or not Path(str(image_path)).is_file():
        return
    legacy_rows = []
    for a in anns:
        bb = a.get("BBox") or []
        if len(bb) < 4:
            continue
        legacy_rows.append(
            {
                "AnnotationType": a.get("AnnotationType") or "Dimensions",
                "BBox": [int(bb[0]), int(bb[1]), int(bb[2]), int(bb[3])],
                "id": a.get("id"),
            }
        )
    if not legacy_rows:
        return
    try:
        pos_map = _tasks().getBalloonCoordinates(str(image_path), legacy_rows)
    except Exception as exc:
        payload["balloon_placement_error"] = str(exc)[:300]
        return
    for a in anns:
        aid = a.get("id")
        if aid is None or aid not in pos_map:
            continue
        tp = pos_map[aid].get("TextPos")
        if tp and len(tp) >= 2:
            a["TextPos"] = [int(tp[0]), int(tp[1])]
    payload["drawing_annotations"] = anns
    payload["balloon_placement"] = "legacy"


def _refine_dimension_detection_payload(payload: dict) -> None:
    """Expand dimension bboxes for vertical/horizontal callouts (internal balloon placement)."""
    if not _dimension_bbox_refine_enabled():
        return
    full = list(payload.get("detections_full") or [])
    if not full:
        return

    infer_path = payload.get("infer_image_path")
    gray = None
    w_full, h_full = 0, 0
    if infer_path and Path(str(infer_path)).is_file():
        gray = cv2.imread(str(infer_path), cv2.IMREAD_GRAYSCALE)
        if gray is not None:
            h_full, w_full = gray.shape[:2]
    if w_full < 1 or h_full < 1:
        w_full = int(payload.get("width") or 0)
        h_full = int(payload.get("height") or 0)

    refined = [_refine_dimension_callout_bbox(d, gray, w_full, h_full) for d in full]
    sc = float(payload.get("width") or w_full) / float(w_full) if w_full > 0 else 1.0
    disp: list = []
    for d in refined:
        bb = d.get("bbox") or []
        if len(bb) < 4:
            continue
        row = {
            "class_name": d.get("class_name"),
            "confidence": d.get("confidence"),
            "bbox": bb if abs(sc - 1.0) < 0.001 else [
                int(bb[0] * sc),
                int(bb[1] * sc),
                int(bb[2] * sc),
                int(bb[3] * sc),
            ],
        }
        if d.get("dimension_orientation"):
            row["dimension_orientation"] = d["dimension_orientation"]
        if d.get("balloon_side"):
            row["balloon_side"] = d["balloon_side"]
        if d.get("source"):
            row["source"] = d["source"]
        disp.append(row)

    payload["detections_full"] = refined
    payload["detections"] = disp or refined
    payload["count"] = len(payload["detections"])
    payload["dimension_bbox_refined"] = True


# Valid YOLO / vision class names (AutoBallooningModel.pt)
_VISION_ALLOWED_CLASSES = {
    "Dimensions",
    "GDnT",
    "Notes",
    "Title_Block",
    "Special_Characteristics",
    "Datums",
    "Welding_Symbols",
    "Surface_Finish_Symbols",
    "Revision_Table",
    "Miscellaneous",
}

_VISION_CLASS_ALIASES = {
    "dimension": "Dimensions",
    "dimensions": "Dimensions",
    "gdt": "GDnT",
    "gd&t": "GDnT",
    "gdn&t": "GDnT",
    "note": "Notes",
    "notes": "Notes",
    "surface finish": "Surface_Finish_Symbols",
    "surface_finish": "Surface_Finish_Symbols",
    "surface finish symbols": "Surface_Finish_Symbols",
    "special characteristic": "Special_Characteristics",
    "special_characteristics": "Special_Characteristics",
    "title block": "Title_Block",
    "title_block": "Title_Block",
    "revision table": "Revision_Table",
    "revision_table": "Revision_Table",
    "revision number": "Revision_Table",
    "part number": "Miscellaneous",
    "part name": "Miscellaneous",
    "mass": "Miscellaneous",
    "weight": "Miscellaneous",
    "mass/weight": "Miscellaneous",
    "finish treatment": "Miscellaneous",
    "finish": "Miscellaneous",
    "date": "Miscellaneous",
    "datum": "Datums",
    "datums": "Datums",
    "welding": "Welding_Symbols",
    "welding symbols": "Welding_Symbols",
    "welding_symbols": "Welding_Symbols",
    "misc": "Miscellaneous",
    "miscellaneous": "Miscellaneous",
}

def _mechanical_ballooning_prompts():
    from Resources.prompts import mechanical_ballooning as mb

    return mb


def _format_yolo_boxes_for_vision_prompt(dets: list, img_w: int, img_h: int) -> str:
    """Compact list of YOLO boxes so Claude does not duplicate existing balloons."""
    lines: list[str] = []
    for i, d in enumerate(dets or [], start=1):
        bb = (d or {}).get("bbox") or []
        if len(bb) < 4:
            continue
        cls = str((d or {}).get("class_name") or "Dimensions")
        cx, cy = _bbox_center(bb)
        gx = int((cx / max(1, img_w)) * 8) + 1 if img_w > 0 else 0
        gy = int((cy / max(1, img_h)) * 6) + 1 if img_h > 0 else 0
        lines.append(
            f"  #{i} {cls} bbox=[{int(bb[0])},{int(bb[1])},{int(bb[2])},{int(bb[3])}] "
            f"grid~col{min(8,gx)}-row{min(6,gy)}"
        )
    if not lines:
        return "  (none)"
    return "\n".join(lines[:80])


def _mark_yolo_detection_sources(payload: dict) -> None:
    for d in payload.get("detections_full") or payload.get("detections") or []:
        if d and not (d.get("source") or "").strip():
            d["source"] = "yolo"


def _region_prepass_enabled() -> bool:
    if _deploy_safe_mode():
        return False
    if not _vision_api_configured():
        return False
    return os.environ.get("BALLOON_REGION_PREPASS", "1").strip().lower() not in (
        "0",
        "false",
        "no",
        "off",
    )


def _opencv_dim_detect_enabled() -> bool:
    if _deploy_safe_mode():
        return False
    return os.environ.get("BALLOON_OPENCV_DIM_LINES", "1").strip().lower() not in (
        "0",
        "false",
        "no",
        "off",
    )


def _ocr_engine() -> str:
    """claude (default when API key) | tesseract. Render safe mode defaults to tesseract (fast, no API)."""
    raw = os.environ.get("BALLOON_OCR_ENGINE", "").strip().lower()
    if raw in ("tesseract", "tess", "local"):
        return "tesseract"
    if raw in ("claude", "anthropic", "vision"):
        return "claude"
    if _deploy_safe_mode():
        return "tesseract"
    return "claude" if _vision_api_configured() else "tesseract"


def _format_opencv_boxes_for_prompt(candidates: list, orig_w: int, orig_h: int, limit: int = 40) -> str:
    lines = []
    for i, d in enumerate((candidates or [])[:limit], start=1):
        bb = (d or {}).get("bbox") or []
        if len(bb) < 4:
            continue
        lines.append(
            f"  {i}. [{int(bb[0])},{int(bb[1])},{int(bb[2])},{int(bb[3])}] "
            f"({orig_w}x{orig_h})"
        )
    return "\n".join(lines) if lines else "(none)"


def _vision_bbox_gap_fill_prompt(
    yolo_dets: list,
    orig_w: int,
    orig_h: int,
    opencv_candidates: list | None = None,
    region_name: str = "",
) -> str:
    cols = int(os.environ.get("BALLOON_VISION_GRID_COLS", "8"))
    rows = int(os.environ.get("BALLOON_VISION_GRID_ROWS", "6"))
    yolo_text = _format_yolo_boxes_for_vision_prompt(yolo_dets, orig_w, orig_h)
    opencv_text = ""
    if opencv_candidates:
        opencv_text = _format_opencv_boxes_for_prompt(opencv_candidates, orig_w, orig_h)
    return _mechanical_ballooning_prompts().anthropic_gap_fill_after_yolo_prompt(
        yolo_text,
        cols,
        rows,
        opencv_candidates_text=opencv_text,
        region_name=region_name,
    )


def _anthropic_region_prepass(payload: dict) -> None:
    """Claude segments the sheet into named view regions before gap-fill."""
    if not _region_prepass_enabled():
        return
    infer_path = payload.get("infer_image_path")
    if not infer_path or not Path(str(infer_path)).is_file():
        return
    try:
        from drawing_regions import parse_regions_from_llm  # type: ignore
    except ImportError:
        return
    max_side = int(os.environ.get("BALLOON_VISION_MAX_SIDE", "2048"))
    try:
        image_bytes, orig_w, orig_h, _ = _prepare_image_bytes_for_vision(str(infer_path), max_side)
    except Exception as exc:
        payload["region_prepass_error"] = str(exc)[:300]
        return
    prompt = (
        _mechanical_ballooning_prompts().anthropic_region_segmentation_prompt()
        + f"\nImage size: {orig_w} x {orig_h} pixels (width x height). "
        "Return coordinates in that pixel space.\n"
    )
    raw = _vision_llm_chat_direct(
        image_bytes,
        prompt,
        max_tokens=int(os.environ.get("BALLOON_REGION_MAX_TOKENS", "4096")),
        temperature=0.05,
    )
    if str(raw).strip().startswith("VISION_LLM_FAILED"):
        payload["region_prepass_error"] = str(raw).strip()[:300]
        return
    parsed = _parse_json_object_from_llm(raw)
    regions = parse_regions_from_llm(parsed, orig_w, orig_h)
    if not regions:
        payload["region_prepass_skipped"] = "no_regions"
        return
    payload["view_regions"] = regions
    payload["region_prepass_count"] = len(regions)
    full = list(payload.get("detections_full") or payload.get("detections") or [])
    try:
        from drawing_regions import tag_detections_with_regions  # type: ignore

        tag_detections_with_regions(full, regions)
    except ImportError:
        pass
    _apply_detections_full_to_payload(payload, full)


def _opencv_dim_line_stage(payload: dict) -> list:
    """Hough extension-line pairs → candidate bboxes; optional direct merge."""
    if not _opencv_dim_detect_enabled():
        return []
    infer_path = payload.get("infer_image_path")
    if not infer_path or not Path(str(infer_path)).is_file():
        return []
    try:
        from dim_line_detect import detect_dim_line_candidates  # type: ignore

        max_c = int(os.environ.get("BALLOON_OPENCV_MAX_CANDIDATES", "80"))
        candidates = detect_dim_line_candidates(str(infer_path), max_candidates=max_c)
    except Exception as exc:
        payload["opencv_dim_error"] = str(exc)[:300]
        return []
    payload["opencv_dim_candidates"] = candidates
    payload["opencv_dim_candidate_count"] = len(candidates)
    if os.environ.get("BALLOON_OPENCV_AUTO_MERGE", "1").strip().lower() in (
        "0",
        "false",
        "no",
        "off",
    ):
        return candidates
    merge_iou = float(os.environ.get("BALLOON_OPENCV_MERGE_IOU", "0.42"))
    full = list(payload.get("detections_full") or payload.get("detections") or [])
    added = 0
    for cand in candidates:
        bb = cand.get("bbox") or []
        if len(bb) < 4:
            continue
        if _max_bbox_iou_with_list(bb, full) >= merge_iou:
            continue
        full.append(cand)
        added += 1
    if added:
        _apply_detections_full_to_payload(payload, full)
        _cleanup_detection_payload(payload)
    payload["opencv_dim_merged"] = added
    return candidates


def _region_crop_jpeg_bytes(image_path: str, region: dict, pad: float = 0.02) -> tuple[bytes, int, int, float] | None:
    """Crop region from drawing; return jpeg bytes and crop offset (ox, oy) and scale."""
    img = _imread_bgr(image_path)
    if img is None:
        return None
    h, w = img.shape[:2]
    bb = region.get("bbox") or []
    if len(bb) < 4:
        return None
    x1, y1, x2, y2 = [int(v) for v in bb[:4]]
    pad_x = int((x2 - x1) * pad)
    pad_y = int((y2 - y1) * pad)
    x1 = max(0, x1 - pad_x)
    y1 = max(0, y1 - pad_y)
    x2 = min(w, x2 + pad_x)
    y2 = min(h, y2 + pad_y)
    crop = img[y1:y2, x1:x2]
    if crop.size < 1:
        return None
    max_side = int(os.environ.get("BALLOON_VISION_MAX_SIDE", "2048"))
    ch, cw = crop.shape[:2]
    scale = 1.0
    if max(ch, cw) > max_side:
        scale = max_side / max(ch, cw)
        crop = cv2.resize(
            crop,
            (max(1, int(round(cw * scale))), max(1, int(round(ch * scale)))),
            interpolation=cv2.INTER_AREA,
        )
    ok, buf = cv2.imencode(".jpg", crop, [int(cv2.IMWRITE_JPEG_QUALITY), 90])
    if not ok:
        return None
    return buf.tobytes(), x1, y1, scale


def _offset_vision_detections_to_full(
    vision_dets: list, ox: int, oy: int, inv_scale: float
) -> list:
    out = []
    for vd in vision_dets or []:
        bb = list((vd or {}).get("bbox") or [])
        if len(bb) < 4:
            continue
        bb = [
            int(round(ox + bb[0] * inv_scale)),
            int(round(oy + bb[1] * inv_scale)),
            int(round(ox + bb[2] * inv_scale)),
            int(round(oy + bb[3] * inv_scale)),
        ]
        nd = dict(vd)
        nd["bbox"] = bb
        nd["region_offset"] = [ox, oy]
        out.append(nd)
    return out


def _deploy_safe_mode() -> bool:
    """
    Render / small instances: skip slow Claude grid passes and cap per-crop OCR
    to avoid HTTP 502 (proxy timeout or OOM). Set BALLOON_RENDER_SAFE=0 to disable.
    """
    explicit = os.environ.get("BALLOON_RENDER_SAFE", "").strip().lower()
    if explicit in ("0", "false", "no", "off"):
        return False
    if explicit in ("1", "true", "yes", "on"):
        return True
    return os.environ.get("RENDER", "").strip().lower() in ("true", "1")


def _max_crop_ocr_count() -> int:
    raw = os.environ.get("BALLOON_MAX_CROP_OCR", "").strip()
    if raw:
        try:
            return max(0, int(raw))
        except ValueError:
            pass
    if _deploy_safe_mode():
        # Claude per-crop OCR is the main cause of Render HTTP 502 (proxy timeout).
        return 6 if _ocr_engine() == "claude" else 80
    return 99999


def _vision_fallback_mode() -> str:
    """together (default) = YOLO first, then always merge Anthropic bboxes. lazy = Anthropic only if sparse."""
    return os.environ.get("BALLOON_VISION_FALLBACK", "together").strip().lower()


def _should_run_anthropic_bbox_supplement(mode: str, yolo_count: int, dim_count: int) -> bool:
    if mode in ("never", "0", "false", "off", "yolo_only", "yolo-only"):
        return False
    if mode in (
        "together",
        "with_yolo",
        "yolo_and_anthropic",
        "supplement",
        "always",
        "auto",
        "1",
        "true",
        "on",
        "all",
    ):
        return True
    if mode in ("lazy", "sparse", "if_sparse"):
        min_yolo = int(os.environ.get("BALLOON_VISION_FALLBACK_MIN_YOLO", "8"))
        min_dims = int(os.environ.get("BALLOON_VISION_FALLBACK_MIN_DIMS", "18"))
        return yolo_count < min_yolo or dim_count < min_dims
    return True


def _anthropic_api_key() -> str:
    key = (os.environ.get("ANTHROPIC_API_KEY") or "").strip()
    if key:
        return key
    try:
        cfg = config.GetConfiguration("ANTHROPIC") or {}
        return (cfg.get("API_KEY") or cfg.get("api_key") or "").strip()
    except Exception:
        return ""


def _openai_api_key() -> str:
    key = (os.environ.get("OPENAI_API_KEY") or "").strip()
    if key:
        return key
    try:
        cfg = config.GetConfiguration("OPENAI") or {}
        return (cfg.get("openai_api_key") or cfg.get("API_KEY") or "").strip()
    except Exception:
        return ""


def _vision_llm_provider() -> str:
    pref = (os.environ.get("BALLOON_VISION_PROVIDER") or "anthropic").strip().lower()
    if pref == "openai" and _openai_api_key():
        return "openai"
    if pref in ("anthropic", "claude") and _anthropic_api_key():
        return "anthropic"
    if _anthropic_api_key():
        return "anthropic"
    if _openai_api_key():
        return "openai"
    return ""


def _balloon_vision_model(provider: str | None = None) -> str:
    model = (os.environ.get("BALLOON_VISION_MODEL") or "").strip()
    if model:
        return model
    prov = provider or _vision_llm_provider()
    try:
        if prov == "anthropic":
            cfg = config.GetConfiguration("ANTHROPIC") or {}
            return (cfg.get("VISION_MODEL") or "claude-sonnet-4-6").strip()
        cfg = config.GetConfiguration("OPENAI") or {}
        return (cfg.get("VISION_MODEL") or "gpt-4o").strip()
    except Exception:
        return "claude-sonnet-4-6" if prov == "anthropic" else "gpt-4o"


def _anthropic_vision_chat_direct(
    image_bytes: bytes, prompt: str, max_tokens: int = 4096, temperature: float = 0.15
) -> str:
    """Call Anthropic Messages API with image (Claude vision)."""
    import base64
    import requests

    key = _anthropic_api_key()
    if not key:
        return "VISION_LLM_FAILED: ANTHROPIC_API_KEY is not set"
    model = _balloon_vision_model("anthropic")
    b64 = base64.b64encode(image_bytes).decode("utf-8")
    headers = {
        "x-api-key": key,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }
    data = {
        "model": model,
        "max_tokens": int(max_tokens),
        "temperature": float(temperature),
        "messages": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": "image/jpeg",
                            "data": b64,
                        },
                    },
                    {"type": "text", "text": prompt},
                ],
            }
        ],
    }
    try:
        response = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers=headers,
            json=data,
            timeout=180,
        )
        res = response.json()
        if response.status_code != 200:
            err = res.get("error", {})
            msg = err.get("message", response.text) if isinstance(err, dict) else str(res)
            return f"VISION_LLM_FAILED: Anthropic HTTP {response.status_code}: {msg}"
        for block in res.get("content") or []:
            if isinstance(block, dict) and block.get("type") == "text":
                return str(block.get("text") or "")
        return f"VISION_LLM_FAILED: {res}"
    except Exception as exc:
        return f"VISION_LLM_FAILED: {exc}"


def _openai_vision_chat_direct(
    image_bytes: bytes, prompt: str, max_tokens: int = 4096, temperature: float = 0.15
) -> str:
    """Call OpenAI vision directly (fallback when Anthropic is not configured)."""
    import base64
    import requests

    key = _openai_api_key()
    if not key:
        return "VISION_LLM_FAILED: OPENAI_API_KEY is not set"
    model = _balloon_vision_model("openai")
    b64 = base64.b64encode(image_bytes).decode("utf-8")
    headers = {"Authorization": f"Bearer {key}", "Content-Type": "application/json"}
    content = [
        {"type": "text", "text": prompt},
        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}"}},
    ]
    data = {
        "model": model,
        "messages": [{"role": "user", "content": content}],
        "max_tokens": int(max_tokens),
        "temperature": float(temperature),
    }
    try:
        response = requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers=headers,
            json=data,
            timeout=180,
        )
        res = response.json()
        if response.status_code != 200:
            err = res.get("error", {})
            msg = err.get("message", response.text) if isinstance(err, dict) else str(res)
            return f"VISION_LLM_FAILED: OpenAI HTTP {response.status_code}: {msg}"
        if res.get("choices"):
            return str(res["choices"][0]["message"]["content"])
        return f"VISION_LLM_FAILED: {res}"
    except Exception as exc:
        return f"VISION_LLM_FAILED: {exc}"


def _vision_llm_chat_direct(
    image_bytes: bytes, prompt: str, max_tokens: int = 4096, temperature: float = 0.15
) -> str:
    """Vision LLM: Anthropic (default) or OpenAI fallback."""
    prov = _vision_llm_provider()
    if prov == "anthropic":
        return _anthropic_vision_chat_direct(image_bytes, prompt, max_tokens, temperature)
    if prov == "openai":
        return _openai_vision_chat_direct(image_bytes, prompt, max_tokens, temperature)
    return "VISION_LLM_FAILED: Set ANTHROPIC_API_KEY (preferred) or OPENAI_API_KEY in backend/.env"


def _vision_api_configured() -> bool:
    return bool(_vision_llm_provider())


def _normalize_vision_class_name(raw: str) -> str:
    s = str(raw or "").strip()
    if not s:
        return "Miscellaneous"
    if s in _VISION_ALLOWED_CLASSES:
        return s
    key = s.lower().replace("-", " ").replace("_", " ")
    if key in _VISION_CLASS_ALIASES:
        return _VISION_CLASS_ALIASES[key]
    compact = key.replace(" ", "_")
    for alias, canonical in _VISION_CLASS_ALIASES.items():
        if compact == alias.replace(" ", "_"):
            return canonical
    if "dimension" in key:
        return "Dimensions"
    if "gdt" in key or "gd&t" in key:
        return "GDnT"
    if "note" in key:
        return "Notes"
    if "datum" in key:
        return "Datums"
    if "weld" in key:
        return "Welding_Symbols"
    if "surface" in key and "finish" in key:
        return "Surface_Finish_Symbols"
    if "revision" in key and "table" in key:
        return "Revision_Table"
    if "title" in key and "block" in key:
        return "Title_Block"
    if "special" in key:
        return "Special_Characteristics"
    return "Miscellaneous"


def _confidence_label_to_float(label: str) -> float:
    m = {"high": 0.88, "medium": 0.72, "low": 0.55}
    return m.get(str(label or "").strip().lower(), 0.65)


def _parse_json_object_from_llm(text: str) -> dict | list | None:
    if not text or str(text).strip().startswith("VISION_LLM_FAILED"):
        return None
    t = str(text).strip()
    if t.startswith("```"):
        t = re.sub(r"^```(?:json)?\s*", "", t, flags=re.IGNORECASE)
        t = re.sub(r"\s*```$", "", t)
    try:
        return json.loads(t)
    except json.JSONDecodeError:
        pass
    m = re.search(r"\{[\s\S]*\}", t)
    if m:
        try:
            return json.loads(m.group(0))
        except json.JSONDecodeError:
            pass
    m = re.search(r"\[[\s\S]*\]", t)
    if m:
        try:
            return json.loads(m.group(0))
        except json.JSONDecodeError:
            pass
    return None


def _prepare_image_bytes_for_vision(image_path: str, max_side: int) -> tuple[bytes, int, int, float]:
    """JPEG bytes for vision API, original (w,h), scale = sent_size / original."""
    bgr = cv2.imread(str(image_path))
    if bgr is None:
        raise ValueError(f"Could not read image: {image_path}")
    orig_h, orig_w = bgr.shape[:2]
    scale = 1.0
    long_side = max(orig_w, orig_h)
    if long_side > max_side:
        scale = max_side / float(long_side)
        new_w = max(1, int(round(orig_w * scale)))
        new_h = max(1, int(round(orig_h * scale)))
        bgr = cv2.resize(bgr, (new_w, new_h), interpolation=cv2.INTER_AREA)
    ok, buf = cv2.imencode(".jpg", bgr, [int(cv2.IMWRITE_JPEG_QUALITY), 86])
    if not ok:
        raise ValueError("Failed to encode image for vision API")
    return buf.tobytes(), orig_w, orig_h, scale


def _vision_bbox_from_entry(entry: dict, sent_scale: float, orig_w: int, orig_h: int) -> list | None:
    """Convert one vision JSON entry to full-resolution [x1,y1,x2,y2]."""
    inv = 1.0 / sent_scale if sent_scale > 1e-6 else 1.0

    def _pick_coord(*keys: str) -> int | None:
        for k in keys:
            if k in entry and entry[k] is not None:
                try:
                    return int(float(entry[k]))
                except (TypeError, ValueError):
                    continue
        return None

    x1 = _pick_coord("x_min", "x1", "left", "xmin")
    y1 = _pick_coord("y_min", "y1", "top", "ymin")
    x2 = _pick_coord("x_max", "x2", "right", "xmax")
    y2 = _pick_coord("y_max", "y2", "bottom", "ymax")
    if x1 is None or y1 is None or x2 is None or y2 is None:
        bb = entry.get("bbox")
        if isinstance(bb, (list, tuple)) and len(bb) >= 4:
            x1, y1, x2, y2 = int(bb[0]), int(bb[1]), int(bb[2]), int(bb[3])
        else:
            return None
    x1, x2 = int(x1 * inv), int(x2 * inv)
    y1, y2 = int(y1 * inv), int(y2 * inv)
    if x2 < x1:
        x1, x2 = x2, x1
    if y2 < y1:
        y1, y2 = y2, y1
    x1 = max(0, min(x1, orig_w - 1))
    y1 = max(0, min(y1, orig_h - 1))
    x2 = max(x1 + 1, min(x2, orig_w))
    y2 = max(y1 + 1, min(y2, orig_h))
    if (x2 - x1) < 6 or (y2 - y1) < 6:
        return None
    return [x1, y1, x2, y2]


def _vision_llm_bbox_detections(
    image_path: str,
    yolo_dets: list | None = None,
    prompt_builder: str | None = None,
    image_bytes_override: bytes | None = None,
    coord_orig_w: int | None = None,
    coord_orig_h: int | None = None,
    coord_sent_scale: float | None = None,
    opencv_candidates: list | None = None,
    region_name: str = "",
) -> list[dict]:
    """Stage 2/3: grid-wise gap fill or QC verify — only callouts not already ballooned."""
    max_side = int(os.environ.get("BALLOON_VISION_MAX_SIDE", "2048"))
    if image_bytes_override is not None:
        image_bytes = image_bytes_override
        orig_w = int(coord_orig_w or 0)
        orig_h = int(coord_orig_h or 0)
        sent_scale = float(coord_sent_scale or 1.0)
    else:
        image_bytes, orig_w, orig_h, sent_scale = _prepare_image_bytes_for_vision(
            image_path, max_side
        )
    if prompt_builder == "coverage_verify":
        cols = int(os.environ.get("BALLOON_VISION_GRID_COLS", "8"))
        rows = int(os.environ.get("BALLOON_VISION_GRID_ROWS", "6"))
        body = _mechanical_ballooning_prompts().anthropic_coverage_verify_prompt(
            _format_yolo_boxes_for_vision_prompt(yolo_dets or [], orig_w, orig_h),
            cols,
            rows,
        )
    else:
        body = _vision_bbox_gap_fill_prompt(
            yolo_dets or [],
            orig_w,
            orig_h,
            opencv_candidates=opencv_candidates,
            region_name=region_name,
        )
    prompt = (
        body
        + f"\nThis image is {int(round(orig_w * sent_scale))} x {int(round(orig_h * sent_scale))} pixels "
        f"(width x height). Return coordinates in that pixel space.\n"
    )
    raw = _vision_llm_chat_direct(
        image_bytes,
        prompt,
        max_tokens=int(os.environ.get("BALLOON_VISION_MAX_TOKENS", "8192")),
        temperature=0.05,
    )
    if str(raw).strip().startswith("VISION_LLM_FAILED"):
        raise RuntimeError(str(raw).strip()[:500])
    parsed = _parse_json_object_from_llm(raw)
    rows = []
    if isinstance(parsed, dict):
        rows = parsed.get("detections") or parsed.get("elements") or parsed.get("boxes") or []
    elif isinstance(parsed, list):
        rows = parsed
    out: list[dict] = []
    for entry in rows or []:
        if not isinstance(entry, dict):
            continue
        bb = _vision_bbox_from_entry(entry, sent_scale, orig_w, orig_h)
        if not bb:
            continue
        cls = _normalize_vision_class_name(entry.get("class_name") or entry.get("class") or "")
        conf = _confidence_label_to_float(entry.get("confidence", "medium"))
        out.append(
            {
                "class_name": cls,
                "confidence": conf,
                "bbox": bb,
                "source": (_vision_llm_provider() or "vision") + "_vision",
                "description": str(entry.get("description") or "")[:200],
            }
        )
    return out


def _max_bbox_iou_with_list(bb: list, dets: list) -> float:
    best = 0.0
    for d in dets:
        ob = (d or {}).get("bbox") or []
        if len(ob) < 4:
            continue
        best = max(best, _bbox_overlap(bb, ob))
    return best


def _apply_detections_full_to_payload(payload: dict, full: list) -> None:
    """Write detections_full and preview-scaled detections[] from full-res boxes."""
    infer_path = payload.get("infer_image_path")
    w_full, h_full = 0, 0
    if infer_path and Path(str(infer_path)).is_file():
        gray = cv2.imread(str(infer_path), cv2.IMREAD_GRAYSCALE)
        if gray is not None:
            h_full, w_full = gray.shape[:2]
    if w_full < 1:
        w_full = int(payload.get("width") or 0)
    if h_full < 1:
        h_full = int(payload.get("height") or 0)
    sc = float(payload.get("width") or w_full) / float(w_full) if w_full > 0 else 1.0
    disp: list = []
    for d in full:
        bb = d.get("bbox") or []
        if len(bb) < 4:
            continue
        if abs(sc - 1.0) < 0.001:
            disp.append(d)
        else:
            disp.append(
                {
                    "class_name": d.get("class_name"),
                    "confidence": d.get("confidence"),
                    "bbox": [int(bb[0] * sc), int(bb[1] * sc), int(bb[2] * sc), int(bb[3] * sc)],
                    "source": d.get("source"),
                }
            )
    payload["detections_full"] = full
    payload["detections"] = disp or full
    payload["count"] = len(payload["detections"])


def _apply_vision_fallback_if_needed(payload: dict) -> None:
    """
    Pipeline stage 2 (after YOLO):
    - YOLO detections are kept; balloons will be built from them first.
    - Anthropic scans the full drawing grid-wise and adds ONLY missed callouts
      (horizontal + vertical dimensions), never duplicating existing YOLO boxes.
    """
    if _deploy_safe_mode():
        payload["vision_fallback_skipped"] = "render_safe_mode"
        payload["detection_pipeline"] = "yolo"
        payload["pipeline_logic"] = "yolo_render_safe"
        return
    mode = _vision_fallback_mode()
    if mode in ("never", "0", "false", "off", "yolo_only", "yolo-only"):
        payload["detection_pipeline"] = "yolo"
        payload["pipeline_logic"] = "yolo_only"
        return
    if not _vision_api_configured():
        payload["vision_fallback_skipped"] = "no_api_key"
        payload["detection_pipeline"] = "yolo"
        payload["pipeline_logic"] = "yolo_only_no_claude_key"
        return

    infer_path = payload.get("infer_image_path")
    if not infer_path or not Path(str(infer_path)).is_file():
        payload["vision_fallback_skipped"] = "no_image"
        payload["detection_pipeline"] = "yolo"
        return

    yolo_full = list(payload.get("detections_full") or payload.get("detections") or [])
    yolo_count = len(yolo_full)
    dim_count = sum(
        1
        for d in yolo_full
        if str((d or {}).get("class_name") or "") in ("Dimensions", "GDnT")
    )
    payload["yolo_balloon_base_count"] = yolo_count
    if not _should_run_anthropic_bbox_supplement(mode, yolo_count, dim_count):
        payload["vision_fallback_skipped"] = "yolo_sufficient"
        payload["vision_fallback_yolo_count"] = yolo_count
        payload["vision_fallback_dim_count"] = dim_count
        payload["detection_pipeline"] = "yolo"
        payload["pipeline_logic"] = "yolo_only_sufficient"
        return

    opencv_cands = payload.get("opencv_dim_candidates") or []
    regions = payload.get("view_regions") or []
    vision_dets: list = []
    try:
        if regions and _region_prepass_enabled():
            try:
                from drawing_regions import detections_in_region  # type: ignore
            except ImportError:
                detections_in_region = None  # type: ignore
            for reg in regions:
                region_yolo = (
                    detections_in_region(yolo_full, reg)
                    if detections_in_region
                    else yolo_full
                )
                crop_pack = _region_crop_jpeg_bytes(str(infer_path), reg)
                if not crop_pack:
                    continue
                cbytes, ox, oy, cscale = crop_pack
                bb = reg.get("bbox") or [0, 0, 0, 0]
                cw = max(1, int(bb[2]) - int(bb[0]))
                ch = max(1, int(bb[3]) - int(bb[1]))
                sent_w = int(round(cw * cscale))
                sent_h = int(round(ch * cscale))
                part = _vision_llm_bbox_detections(
                    str(infer_path),
                    yolo_dets=region_yolo,
                    image_bytes_override=cbytes,
                    coord_orig_w=sent_w,
                    coord_orig_h=sent_h,
                    coord_sent_scale=cscale,
                    opencv_candidates=opencv_cands,
                    region_name=str(reg.get("name") or ""),
                )
                vision_dets.extend(
                    _offset_vision_detections_to_full(part, ox, oy, 1.0 / cscale)
                )
        else:
            vision_dets = _vision_llm_bbox_detections(
                str(infer_path),
                yolo_dets=yolo_full,
                opencv_candidates=opencv_cands,
            )
        payload["vision_fallback_provider"] = _vision_llm_provider()
    except Exception as exc:
        payload["vision_fallback_error"] = str(exc)[:500]
        payload["detection_pipeline"] = "yolo"
        payload["pipeline_logic"] = "yolo_claude_gap_fill_failed"
        return

    merge_iou = float(os.environ.get("BALLOON_VISION_MERGE_IOU", "0.40"))
    min_conf = float(os.environ.get("BALLOON_VISION_MIN_CONF", "0.62"))
    full = list(yolo_full)
    added = 0
    for vd in vision_dets:
        bb = vd.get("bbox") or []
        if len(bb) < 4:
            continue
        if float((vd or {}).get("confidence") or 0) < min_conf:
            continue
        if _max_bbox_iou_with_list(bb, yolo_full) >= merge_iou:
            continue
        if _max_bbox_iou_with_list(bb, full) >= merge_iou:
            continue
        vd["source"] = (_vision_llm_provider() or "anthropic") + "_gap_fill"
        full.append(vd)
        added += 1

    _apply_detections_full_to_payload(payload, full)
    _cleanup_detection_payload(payload)
    full = list(payload.get("detections_full") or [])

    payload["vision_fallback_used"] = True
    payload["vision_fallback_added"] = added
    payload["vision_fallback_yolo_count"] = yolo_count
    payload["vision_fallback_dim_count"] = dim_count
    payload["vision_fallback_vision_raw"] = len(vision_dets)
    pipe_tag = "yolo_then_anthropic_gap_fill"
    if regions:
        pipe_tag = "yolo_regions_then_anthropic_gap_fill"
    payload["detection_pipeline"] = pipe_tag
    payload["pipeline_logic"] = (
        f"yolo_{yolo_count}_balloons_then_claude_grid_gap_fill_+{added}"
    )
    payload["yolo_detection_count"] = yolo_count


def _coverage_verify_enabled() -> bool:
    if _deploy_safe_mode():
        return False
    return os.environ.get("BALLOON_COVERAGE_VERIFY", "1").strip().lower() not in (
        "0",
        "false",
        "no",
        "off",
    )


def _anthropic_coverage_verify(payload: dict) -> None:
    """Stage 3 QC agent: Claude finds dimensions still missing after YOLO + gap-fill."""
    if not _coverage_verify_enabled() or not _vision_api_configured():
        return
    infer_path = payload.get("infer_image_path")
    if not infer_path or not Path(str(infer_path)).is_file():
        return
    full = list(payload.get("detections_full") or payload.get("detections") or [])
    if not full:
        return
    try:
        verify_dets = _vision_llm_bbox_detections(
            str(infer_path),
            yolo_dets=full,
            prompt_builder="coverage_verify",
        )
    except Exception as exc:
        payload["coverage_verify_error"] = str(exc)[:300]
        return
    merge_iou = float(os.environ.get("BALLOON_VISION_MERGE_IOU", "0.38"))
    min_conf = float(os.environ.get("BALLOON_VISION_MIN_CONF", "0.60"))
    added = 0
    for vd in verify_dets:
        bb = vd.get("bbox") or []
        if len(bb) < 4:
            continue
        if float((vd or {}).get("confidence") or 0) < min_conf:
            continue
        if _max_bbox_iou_with_list(bb, full) >= merge_iou:
            continue
        vd["source"] = (_vision_llm_provider() or "anthropic") + "_coverage_verify"
        full.append(vd)
        added += 1
    _apply_detections_full_to_payload(payload, full)
    payload["coverage_verify_added"] = added
    payload["coverage_verify_raw"] = len(verify_dets)


def _prune_overlapping_duplicate_detections(payload: dict) -> None:
    """Remove extra table rows from duplicate YOLO boxes on one nX callout."""
    items = list(payload.get("balloon_items") or [])
    if not items:
        return
    primary_det_by_parent: dict[str, int] = {}
    for it in items:
        if not _is_sub_balloon_item(it):
            continue
        p = _parent_balloon_number(it)
        di = _detection_index_for_item(payload, it)
        if not p or di is None:
            continue
        if p not in primary_det_by_parent or di < primary_det_by_parent[p]:
            primary_det_by_parent[p] = int(di)
    if not primary_det_by_parent:
        return

    def row_for_di(di: int) -> dict:
        for it in items:
            if _detection_index_for_item(payload, it) == di:
                return it
        return items[di] if di < len(items) else {}

    kept: list = []
    for it in items:
        if _is_sub_balloon_item(it):
            kept.append(it)
            continue
        di = _detection_index_for_item(payload, it)
        if di is None:
            kept.append(it)
            continue
        this_bb = _bbox_for_balloon_row(payload, di, it)
        drop = False
        for p, primary_di in primary_det_by_parent.items():
            if di == primary_di:
                drop = True
                break
            primary_bb = _bbox_for_balloon_row(payload, primary_di, row_for_di(primary_di))
            if primary_bb and this_bb and _bbox_near_multiplier_duplicate(primary_bb, this_bb):
                drop = True
                break
        if not drop:
            kept.append(it)
    payload["balloon_items"] = kept


def _ensure_drawing_annotations(payload: dict) -> None:
    dets = list(payload.get("detections") or [])
    if not dets:
        return
    anns = list(payload.get("drawing_annotations") or [])
    if len(anns) == len(dets):
        return
    payload["drawing_annotations"] = _drawing_annotations_from_detections(dets)


def _repair_multiplier_drawing_annotations(payload: dict) -> None:
    """Align primary nX detection balloon id with parent number; keep server tblr ids elsewhere."""
    _ensure_drawing_annotations(payload)
    dets = list(payload.get("detections") or [])
    anns = list(payload.get("drawing_annotations") or [])
    items = list(payload.get("balloon_items") or [])
    if not dets:
        return

    primary_det_by_parent: dict[str, int] = {}
    for it in items:
        if not _is_sub_balloon_item(it):
            continue
        p = _parent_balloon_number(it)
        di = _detection_index_for_item(payload, it)
        if not p or di is None:
            continue
        if p not in primary_det_by_parent or di < primary_det_by_parent[p]:
            primary_det_by_parent[p] = int(di)

    def item_for_di(di: int) -> dict:
        for it in items:
            if _detection_index_for_item(payload, it) == di:
                return it
        return items[di] if di < len(items) else {}

    # One balloon per detection. Hide only duplicate YOLO slots on the same nX callout (not all overlaps).
    preserve_legacy_pos = payload.get("balloon_placement") == "legacy"
    new_anns: list = []
    for i, d in enumerate(dets):
        bb = (d or {}).get("bbox") or []
        if i < len(anns):
            row = dict(anns[i])
        else:
            row = {}
        row.pop("draw_suppress", None)
        row.pop("report_only", None)
        row.pop("is_parent_balloon", None)
        pid = row.get("id", i + 1)
        canvas_skip = False
        this_bb = _bbox_for_balloon_row(payload, i, item_for_di(i))
        if primary_det_by_parent and this_bb:
            for p, primary_di in primary_det_by_parent.items():
                if i == primary_di:
                    pid = p
                    break
                primary_bb = _bbox_for_balloon_row(payload, primary_di, item_for_di(primary_di))
                if primary_bb and _bbox_near_multiplier_duplicate(primary_bb, this_bb):
                    canvas_skip = True
                    break
        row["id"] = pid
        row["display_id"] = pid
        row["canvas_skip"] = canvas_skip
        if len(bb) >= 4:
            x1, y1, x2, y2 = bb[0], bb[1], bb[2], bb[3]
            row["AnnotationType"] = (d or {}).get("class_name") or row.get("AnnotationType") or "Dimensions"
            row["BBox"] = [int(x1), int(y1), int(x2), int(y2)]
            tp = row.get("TextPos") or []
            if not (
                preserve_legacy_pos
                and isinstance(tp, (list, tuple))
                and len(tp) >= 2
                and tp[0] is not None
                and tp[1] is not None
            ):
                row["TextPos"] = _tight_balloon_text_pos(
                    bb,
                    orientation=(d or {}).get("dimension_orientation"),
                    balloon_side=(d or {}).get("balloon_side"),
                )
        new_anns.append(row)
    payload["drawing_annotations"] = new_anns
    payload.pop("canvas_balloon_annotations", None)


def _expand_multiplier_balloons_payload(payload: dict) -> None:
    """nX: split balloon_items only; one drawing balloon per detection."""
    items = list(payload.get("balloon_items") or [])
    if not items:
        return
    if any(_is_sub_balloon_item(it) for it in items):
        _prune_parent_balloons_with_subs(payload)
        _repair_multiplier_drawing_annotations(payload)
        _sync_balloon_items_from_detections(payload)
        return

    new_items: list = []
    for i, it in enumerate(items):
        if _is_sub_balloon_item(it):
            continue
        row = dict(it)
        det_idx = row.get("detection_index")
        if det_idx is None:
            det_idx = i
        else:
            det_idx = int(det_idx)
        row["detection_index"] = det_idx
        parent_num = row.get("balloon_number", det_idx + 1)
        mult = _multiplier_count_from_item(row)
        if mult >= 2:
            row["multiplier_count"] = mult
            row["multiplier_notation"] = _multiplier_notation(_balloon_item_text(row), mult)
            for k in range(1, mult + 1):
                new_items.append(
                    {
                        **row,
                        "balloon_number": f"{parent_num}.{k}",
                        "parent_balloon_number": parent_num,
                        "sub_balloon_index": k,
                        "is_sub_balloon": True,
                        "is_parent_balloon": False,
                        "detection_index": det_idx,
                    }
                )
            continue
        row["balloon_number"] = parent_num
        new_items.append(row)

    payload["balloon_items"] = new_items
    _prune_parent_balloons_with_subs(payload)
    _repair_multiplier_drawing_annotations(payload)
    _sync_balloon_items_from_detections(payload)


def _tight_balloon_text_pos(
    bb: list,
    orientation: str | None = None,
    balloon_side: str | None = None,
) -> list:
    """Place balloon beside dimension: below horizontal, left/right of vertical length dims."""
    if not bb or len(bb) < 4:
        return []
    x1, y1, x2, y2 = float(bb[0]), float(bb[1]), float(bb[2]), float(bb[3])
    w, h = x2 - x1, y2 - y1
    cx, cy = (x1 + x2) / 2, (y1 + y2) / 2
    gap = max(18.0, min(w, h) * 0.18, max(w, h) * 0.08)
    ori = orientation or _dimension_callout_orientation(w, h)
    if ori == "vertical" or (ori != "horizontal" and h >= w * 1.15):
        side = (balloon_side or "right").lower()
        if side == "left":
            return [x1 - gap, cy]
        return [x2 + gap, cy]
    if ori == "horizontal" or w >= h * 1.15:
        side = (balloon_side or "below").lower()
        if side == "above":
            return [cx, y1 - gap]
        return [cx, y2 + gap]
    return [cx, y2 + gap]


def _drawing_annotations_from_detections(detections: list) -> list:
    out = []
    for i, d in enumerate(detections or [], start=1):
        bb = (d or {}).get("bbox") or []
        if len(bb) >= 4:
            x1, y1, x2, y2 = bb[0], bb[1], bb[2], bb[3]
            tp = _tight_balloon_text_pos(
                bb,
                orientation=(d or {}).get("dimension_orientation"),
                balloon_side=(d or {}).get("balloon_side"),
            )
            row = {
                "id": i,
                "AnnotationType": (d or {}).get("class_name") or "Dimensions",
                "BBox": [int(x1), int(y1), int(x2), int(y2)],
                "TextPos": [int(tp[0]), int(tp[1])] if tp else [],
            }
            if (d or {}).get("dimension_orientation"):
                row["dimension_orientation"] = d["dimension_orientation"]
            if (d or {}).get("balloon_side"):
                row["balloon_side"] = d["balloon_side"]
            out.append(row)
        else:
            out.append(
                {
                    "id": i,
                    "AnnotationType": (d or {}).get("class_name") or "Dimensions",
                    "BBox": [],
                    "TextPos": [],
                }
            )
    return out


def _imread_bgr(path: str):
    """
    Load BGR image from disk. Uses imdecode + fromfile so Unicode paths work on Windows
    (cv2.imread often fails for non-ASCII paths and can mis-read some PNG modes).
    """
    p = Path(path)
    if not p.is_file():
        return None
    try:
        raw = np.fromfile(str(p), dtype=np.uint8)
        img = cv2.imdecode(raw, cv2.IMREAD_COLOR)
        if img is None:
            img = cv2.imdecode(raw, cv2.IMREAD_UNCHANGED)
            if img is not None and len(img.shape) == 2:
                img = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)
            elif img is not None and len(img.shape) == 3 and img.shape[2] == 4:
                img = cv2.cvtColor(img, cv2.COLOR_BGRA2BGR)
        if img is not None:
            return img
    except Exception:
        pass
    return cv2.imread(str(p))


def _balloon_vision_prompt(class_name: str) -> str:
    return _mechanical_ballooning_prompts().crop_extraction_prompt(
        (class_name or "unknown").strip()
    )


def _parse_extraction_json(text: str) -> dict[str, str]:
    """Parse engineer-prompt fields from vision model output."""
    out = {
        "nominal_value": "",
        "tolerance": "",
        "others": "",
        "feature_type": "",
        "view_location": "",
        "inspection_method": "",
        "remarks": "",
    }
    if not text or not str(text).strip():
        return out
    t = str(text).strip()
    if t.startswith("VISION_LLM_FAILED"):
        return out
    if t.startswith("```"):
        t = re.sub(r"^```(?:json)?\s*", "", t, flags=re.IGNORECASE)
        t = re.sub(r"\s*```$", "", t)
    try:
        data = json.loads(t)
        if isinstance(data, dict):
            for key in out:
                if key in data:
                    out[key] = str(data.get(key, "") or "").strip()
            dim = str(data.get("dimension_callout") or "").strip()
            if dim and not out["nominal_value"]:
                out["nominal_value"] = dim
            return out
    except json.JSONDecodeError:
        pass
    out["others"] = t[:2000]
    return out


def _rotate_bgr_k(bgr: np.ndarray, k: int) -> np.ndarray:
    """Rotate BGR by k*90° clockwise (k in 0..3)."""
    if bgr is None or not getattr(bgr, "size", 0) or k % 4 == 0:
        return bgr
    if k % 4 == 1:
        return cv2.rotate(bgr, cv2.ROTATE_90_CLOCKWISE)
    if k % 4 == 2:
        return cv2.rotate(bgr, cv2.ROTATE_180)
    return cv2.rotate(bgr, cv2.ROTATE_90_COUNTERCLOCKWISE)


def _ocr_try_all_rotations_enabled() -> bool:
    return os.environ.get("BALLOON_OCR_TRY_ALL_ROTATIONS", "1").strip().lower() not in (
        "0",
        "false",
        "no",
        "off",
    )


def _ink_rotation_k_hint(bgr: np.ndarray) -> int | None:
    """Guess 90° rotation from ink layout (vertical stroke + sideways digits)."""
    if bgr is None or not getattr(bgr, "size", 0):
        return None
    gray = cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY)
    _, th = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    h, w = th.shape[:2]
    if h >= w * 1.2:
        return 1
    if w >= h * 1.2:
        return 0
    try:
        n_labels, _labels, stats, _ = cv2.connectedComponentsWithStats(th, connectivity=8)
    except Exception:
        n_labels = 0
    best_area = 0
    best_rw, best_rh = 0.0, 0.0
    for i in range(1, n_labels):
        area = int(stats[i, cv2.CC_STAT_AREA])
        if area < 25 or area < best_area:
            continue
        x = int(stats[i, cv2.CC_STAT_LEFT])
        y = int(stats[i, cv2.CC_STAT_TOP])
        bw = int(stats[i, cv2.CC_STAT_WIDTH])
        bh = int(stats[i, cv2.CC_STAT_HEIGHT])
        if bw < 2 or bh < 2:
            continue
        best_area = area
        best_rw, best_rh = float(bw), float(bh)
    if best_area > 0:
        if best_rh >= best_rw * 1.25:
            return 1
        if best_rw >= best_rh * 1.25:
            return 0
    coords = np.column_stack(np.where(th > 0))
    if len(coords) < 12:
        return None
    pts = coords[:, ::-1].astype(np.float32)
    rect = cv2.minAreaRect(pts)
    rw, rh = rect[1]
    if rw < 2 or rh < 2:
        return None
    if rh >= rw * 1.25:
        return 1
    if rw >= rh * 1.25:
        return 0
    return None


def _pick_best_rotation_k(bgr: np.ndarray, orientation: str | None = None) -> int:
    """Pick 0..3 (×90° CW) with strongest Tesseract signal; ink/orientation hints if OCR unavailable."""
    if bgr is None or not getattr(bgr, "size", 0):
        return 0
    h, w = bgr.shape[:2]
    ori = orientation or _dimension_callout_orientation(float(w), float(h))
    ink_k = _ink_rotation_k_hint(bgr)
    order = [0, 1, 3, 2]
    if ori == "vertical":
        order = [1, 3, 0, 2]
    if ink_k is not None and ink_k not in order:
        order.insert(0, ink_k)
    elif ink_k is not None:
        order.remove(ink_k)
        order.insert(0, ink_k)
    try_all = _ocr_try_all_rotations_enabled()
    if not try_all and ori != "vertical" and h < w * 1.08 and w < h * 1.08:
        return 0
    best_k, best_sc = 0, -1.0
    for k in order:
        text = _ocr_text_from_bgr(_rotate_bgr_k(bgr, k))
        sc = _ocr_text_quality_score(text)
        if sc > best_sc:
            best_sc, best_k = sc, k
    if best_sc > 0:
        return best_k
    if ori == "vertical" or (h >= w * 1.05 and ink_k in (1, 3)):
        return ink_k if ink_k is not None else 1
    if ink_k is not None:
        return ink_k
    return 0


def _upscale_bgr_for_ocr(bgr: np.ndarray, min_side: int = 120, target_max: int = 420) -> np.ndarray:
    """Upscale small crops so OCR can read thin/rotated dimension text."""
    if bgr is None or not getattr(bgr, "size", 0):
        return bgr
    h, w = bgr.shape[:2]
    scale = 1.0
    if min(h, w) < min_side:
        scale = max(scale, min_side / min(h, w))
    if max(h, w) < target_max:
        scale = max(scale, target_max / max(h, w))
    scale = min(scale, 4.0)
    if scale > 1.01:
        bgr = cv2.resize(
            bgr,
            (max(1, int(round(w * scale))), max(1, int(round(h * scale)))),
            interpolation=cv2.INTER_CUBIC,
        )
    return bgr


def _deskew_crop_for_vision(bgr: np.ndarray, orientation: str | None = None) -> np.ndarray:
    """Rotate crop so dimension digits read horizontally for Claude vision OCR."""
    if bgr is None or not getattr(bgr, "size", 0):
        return bgr
    base = _upscale_bgr_for_ocr(bgr)
    k = _pick_best_rotation_k(base, orientation)
    return _rotate_bgr_k(base, k)


def _ocr_bgr_variants(bgr: np.ndarray, orientation: str | None = None) -> list[np.ndarray]:
    """Original + 90°/180°/270° views so vertical dims (e.g. 12, 38) are readable."""
    if bgr is None or not getattr(bgr, "size", 0):
        return []
    h, w = bgr.shape[:2]
    ori = orientation or _dimension_callout_orientation(float(w), float(h))
    base = _upscale_bgr_for_ocr(bgr)
    variants: list[np.ndarray] = []
    seen: set[tuple[int, int]] = set()

    def add(img: np.ndarray) -> None:
        if img is None or not getattr(img, "size", 0):
            return
        key = (int(img.shape[0]), int(img.shape[1]))
        if key in seen and len(variants) >= 1:
            return
        seen.add(key)
        variants.append(img)

    add(base)
    try_all = _ocr_try_all_rotations_enabled()
    if try_all or ori == "vertical" or h >= w * 1.05 or w >= h * 1.05:
        for k in (1, 2, 3):
            add(_rotate_bgr_k(base, k))
    elif h >= w * 1.12:
        add(_rotate_bgr_k(base, 1))
        add(_rotate_bgr_k(base, 3))
    return variants


def _ocr_text_quality_score(text: str) -> float:
    t = (text or "").strip()
    if not t:
        return 0.0
    digits = re.findall(r"\d+\.?\d*", t)
    score = len(t) * 0.05 + len(digits) * 8.0
    if re.search(r"[ØøΦφRr°±]", t):
        score += 3.0
    if len(t) <= 2 and not digits:
        score *= 0.15
    if re.fullmatch(r"\d{1,4}(?:\.\d+)?", t):
        score += 20.0
    elif digits and len(digits[0]) >= 2:
        score += 12.0
    if len(t) == 1 and t.isalpha():
        score *= 0.2
    return score


def _preprocess_bgr_for_ocr(bgr: np.ndarray) -> np.ndarray:
    """Upscale and binarize a crop for Tesseract OCR."""
    if bgr is None or not getattr(bgr, "size", 0):
        return bgr
    bgr = _upscale_bgr_for_ocr(bgr)
    h, w = bgr.shape[:2]
    gray = cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY)
    gray = cv2.bilateralFilter(gray, 5, 50, 50)
    _, th = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    if float(np.mean(th)) < 127:
        th = cv2.bitwise_not(th)
    return th


def _tesseract_ocr_best(bgr: np.ndarray, orientation: str | None = None) -> str:
    """Try multiple orientations; pick text with strongest dimension signal."""
    best = ""
    best_score = 0.0
    for variant in _ocr_bgr_variants(bgr, orientation):
        text = _ocr_text_from_bgr(variant)
        sc = _ocr_text_quality_score(text)
        if sc > best_score:
            best_score = sc
            best = text
    return best


def _tesseract_fallback_enabled() -> bool:
    return os.environ.get("BALLOON_OCR_TESSERACT_FALLBACK", "1").strip().lower() not in (
        "0",
        "false",
        "no",
        "off",
    )


def _is_weak_dimension_extraction(parsed: dict[str, str]) -> bool:
    """True when nominal is empty or clearly not a dimension (e.g. single letter 'h')."""
    nom = (parsed.get("nominal_value") or "").strip()
    tol = (parsed.get("tolerance") or "").strip()
    if not nom and not tol:
        return True
    if nom and re.search(r"\d", nom):
        return False
    if tol and re.search(r"\d", tol):
        return False
    if len(nom) <= 2 and not re.search(r"[ØøΦφR°]", nom, re.I):
        return True
    return not nom


def _merge_vision_with_local_ocr(
    vision: dict[str, str], bgr: np.ndarray, class_name: str, orientation: str | None = None
) -> dict[str, str]:
    """Fill gaps / fix wrong reads using Tesseract on original + rotated crops."""
    out = dict(vision)
    if not _tesseract_fallback_enabled():
        _inject_multiplier_from_ocr(out, bgr, orientation)
        return out
    tess = _tesseract_ocr_best(bgr, orientation)
    if tess:
        existing_raw = (out.get("raw_ocr") or "").strip()
        out["raw_ocr"] = (existing_raw + " " + tess).strip()[:2000]
        hint = _parse_dimension_text(tess)
        if _is_weak_dimension_extraction(out):
            if hint["nominal_value"]:
                out["nominal_value"] = hint["nominal_value"]
            if hint["tolerance"]:
                out["tolerance"] = hint["tolerance"]
            if not (out.get("others") or "").strip():
                out["others"] = tess[:2000]
        elif hint["nominal_value"] and re.search(r"\d", hint["nominal_value"]):
            nom = (out.get("nominal_value") or "").strip()
            if not re.search(r"\d", nom) or not _llm_value_supported_by_ocr(nom, tess):
                out["nominal_value"] = hint["nominal_value"]
                if hint["tolerance"]:
                    out["tolerance"] = hint["tolerance"]
    _inject_multiplier_from_ocr(out, bgr, orientation)
    cls = (class_name or "").lower()
    if "gdt" in cls or "g&t" in cls or "gd" in cls:
        if not out["nominal_value"] and out["tolerance"]:
            out["nominal_value"], out["tolerance"] = out["tolerance"], ""
    return out


def _configure_tesseract() -> bool:
    """Point pytesseract at the Windows installer path when not on PATH."""
    try:
        import pytesseract  # type: ignore
    except ImportError:
        return False
    if os.environ.get("TESSERACT_CMD", "").strip():
        pytesseract.pytesseract.tesseract_cmd = os.environ["TESSERACT_CMD"].strip()
        return True
    for cand in (
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    ):
        if os.path.isfile(cand):
            pytesseract.pytesseract.tesseract_cmd = cand
            return True
    return True


def _claude_vision_ocr_crop(
    bgr: np.ndarray, class_name: str, orientation: str | None = None
) -> dict[str, str]:
    """Claude vision on deskewed crop; Tesseract fallback fills vertical / missed numbers."""
    empty = {
        "nominal_value": "",
        "tolerance": "",
        "others": "",
        "feature_type": "",
        "view_location": "",
        "inspection_method": "",
        "remarks": "",
        "raw_ocr": "",
    }
    if bgr is None or not getattr(bgr, "size", 0):
        return _merge_vision_with_local_ocr(empty, bgr, class_name, orientation)
    vision = dict(empty)

    def _vision_from_bgr(send_bgr: np.ndarray) -> dict[str, str]:
        out = dict(empty)
        ok, buf = cv2.imencode(".jpg", send_bgr, [int(cv2.IMWRITE_JPEG_QUALITY), 92])
        if not ok:
            return out
        try:
            val = _tasks()._vision_llm_message(
                buf.tobytes(),
                _mechanical_ballooning_prompts().crop_extraction_prompt(
                    class_name, orientation or ""
                ),
                max_tokens=int(os.environ.get("BALLOON_CROP_MAX_TOKENS", "800")),
                temperature=0.0,
                top_p=1.0,
            )
            parsed = _parse_extraction_json(val or "")
            out["nominal_value"] = (parsed.get("nominal_value") or "").strip()
            out["tolerance"] = (parsed.get("tolerance") or "").strip()
            out["others"] = (parsed.get("others") or "").strip()[:2000]
            out["feature_type"] = (parsed.get("feature_type") or "").strip()
            out["view_location"] = (parsed.get("view_location") or "").strip()
            out["inspection_method"] = (parsed.get("inspection_method") or "").strip()
            out["remarks"] = (parsed.get("remarks") or "").strip()
            raw = " ".join(
                x
                for x in (out["nominal_value"], out["tolerance"], out["others"])
                if x
            ).strip()
            out["raw_ocr"] = raw[:2000]
        except Exception:
            pass
        return out

    if _vision_api_configured():
        base = _upscale_bgr_for_ocr(bgr)
        primary_k = _pick_best_rotation_k(base, orientation)
        send = _rotate_bgr_k(base, primary_k)
        vision = _vision_from_bgr(send)
        if _is_weak_dimension_extraction(vision):
            h, w = base.shape[:2]
            ori = orientation or _dimension_callout_orientation(float(w), float(h))
            alt_ks = [1, 3, 0, 2] if ori == "vertical" else [1, 3, 2]
            for k in alt_ks:
                if k == primary_k:
                    continue
                alt = _vision_from_bgr(_rotate_bgr_k(base, k))
                if not _is_weak_dimension_extraction(alt):
                    vision = alt
                    break
    return _merge_vision_with_local_ocr(vision, bgr, class_name, orientation)


def _ocr_text_from_bgr(bgr: np.ndarray) -> str:
    """Local OCR fallback when vision LLM keys are not configured."""
    if bgr is None or not getattr(bgr, "size", 0):
        return ""
    try:
        import pytesseract  # type: ignore
    except ImportError:
        return ""
    if not _configure_tesseract():
        return ""
    try:
        proc = _preprocess_bgr_for_ocr(bgr)
        ph, pw = proc.shape[:2]
        small = max(ph, pw) < 110
        if small:
            psm_order = (10, 7, 8, 6, 13, 11)
        elif ph >= pw * 1.05:
            psm_order = (7, 8, 6, 10, 5, 11)
        else:
            psm_order = (6, 7, 8, 11, 10)
        best = ""
        best_sc = 0.0
        for psm in psm_order:
            cfg = f"--psm {psm}"
            text = pytesseract.image_to_string(proc, config=cfg)
            text = re.sub(r"\s+", " ", (text or "").strip())
            sc = _ocr_text_quality_score(text)
            if sc > best_sc:
                best_sc, best = sc, text
            if text and re.search(r"\d", text) and sc >= 8.0:
                return text
        return best
    except Exception:
        return ""


def _parse_dimension_text(raw: str) -> dict[str, str]:
    """Parse engineering dimension / GD&T text into nominal_value and tolerance strings."""
    text = (raw or "").strip()
    out = {"nominal_value": "", "tolerance": ""}
    if not text:
        return out
    s = re.sub(r"^[ØøΦφ]\s*", "", text, flags=re.IGNORECASE)
    s = re.sub(r"\s+", " ", s)

    m = re.match(r"^([+-]?\d+\.?\d*)\s*[±]\s*(\d+\.?\d*)", s)
    if m:
        out["nominal_value"] = m.group(1)
        out["tolerance"] = f"±{m.group(2)}"
        return out

    m = re.match(r"^([+-]?\d+\.?\d*)\s*([+-]\d+\.?\d*)\s*/\s*([+-]\d+\.?\d*)", s)
    if m:
        out["nominal_value"] = m.group(1)
        out["tolerance"] = f"{m.group(2)}/{m.group(3)}"
        return out

    m = re.match(r"^([+-]?\d+\.?\d*)\s*\+\s*(\d+\.?\d*)\s*/\s*-?\s*(\d+\.?\d*)", s)
    if m:
        out["nominal_value"] = m.group(1)
        out["tolerance"] = f"+{m.group(2)}/-{m.group(3).lstrip('-')}"
        return out

    m = re.match(r"^([+-]?\d+\.?\d*)\s*([+-]\d+\.?\d*)\s+([+-]\d+\.?\d*)", s)
    if m:
        out["nominal_value"] = m.group(1)
        out["tolerance"] = f"{m.group(2)} {m.group(3)}"
        return out

    # Range: 7.9 - 8.1 or 7.9-8.1
    m = re.search(
        r"([+-]?\d+\.?\d*)\s*[-–]\s*([+-]?\d+\.?\d*)",
        s,
    )
    if m:
        out["nominal_value"] = str((float(m.group(1)) + float(m.group(2))) / 2)
        half = abs(float(m.group(2)) - float(m.group(1))) / 2
        out["tolerance"] = f"±{half:g}"
        return out

    m = re.match(r"^([+-]?\d+\.?\d*)$", s)
    if m:
        out["nominal_value"] = m.group(1)
        return out

    # GD&T / composite: take first numeric as nominal, rest as tolerance
    nums = re.findall(r"[+-]?\d+\.?\d*", s)
    if nums:
        out["nominal_value"] = nums[0]
        if len(nums) > 1:
            out["tolerance"] = " ".join(nums[1:])
        elif re.search(r"[±/]", s):
            out["tolerance"] = s
    return out


def _inject_multiplier_from_ocr(
    out: dict[str, str], bgr: np.ndarray, orientation: str | None = None
) -> None:
    """If vision omitted 2X/3X/nX, recover quantity prefix from Tesseract on the crop."""
    combined = " ".join(
        p
        for p in (out.get("others"), out.get("nominal_value"), out.get("tolerance"))
        if p
    )
    if _parse_multiplier_count(combined) >= 2:
        return
    ocr_text = _tesseract_ocr_best(bgr, orientation)
    if not ocr_text:
        ocr_text = _ocr_text_from_bgr(bgr)
    if not ocr_text:
        return
    out["raw_ocr"] = ocr_text[:2000]
    m = re.search(r"(\d+)\s*[xX×]", ocr_text)
    if not m:
        return
    prefix = re.sub(r"\s+", "", m.group(0))
    others = (out.get("others") or "").strip()
    compact = re.sub(r"\s+", "", others).lower()
    if prefix.lower() not in compact:
        out["others"] = (prefix + (" " + others if others else "")).strip()[:2000]


def _normalize_ocr_blob(text: str) -> str:
    return re.sub(r"[^0-9A-Za-zØøΦφ±°.,/×xX\-\s]", "", (text or "")).lower()


def _llm_value_supported_by_ocr(value: str, ocr_text: str) -> bool:
    """Reject hallucinated LLM values not present in OCR."""
    v = (value or "").strip()
    if not v:
        return True
    ocr = _normalize_ocr_blob(ocr_text)
    if not ocr:
        return False
    vn = _normalize_ocr_blob(v)
    if vn and vn in ocr:
        return True
    digits = re.findall(r"\d+\.?\d*", v)
    if digits and all(d in ocr for d in digits):
        return True
    return False


def _ocr_first_parse_crop(
    bgr: np.ndarray, class_name: str, orientation: str | None = None
) -> dict[str, str]:
    """Claude vision + Tesseract fallback (rotated crops for vertical dims)."""
    if _ocr_engine() == "claude" and bgr is not None:
        return _claude_vision_ocr_crop(bgr, class_name, orientation)
    ocr_text = _tesseract_ocr_best(bgr, orientation) if bgr is not None else ""
    out = {
        "nominal_value": "",
        "tolerance": "",
        "others": (ocr_text or "")[:2000],
        "feature_type": "",
        "view_location": "",
        "inspection_method": "",
        "remarks": "",
        "raw_ocr": (ocr_text or "")[:2000],
    }
    if ocr_text:
        hint = _parse_dimension_text(ocr_text)
        out["nominal_value"] = hint["nominal_value"]
        out["tolerance"] = hint["tolerance"]
    _inject_multiplier_from_ocr(out, bgr, orientation) if bgr is not None else None
    cls = (class_name or "").lower()
    if "gdt" in cls or "g&t" in cls or "gd" in cls:
        if not out["nominal_value"] and out["tolerance"]:
            out["nominal_value"], out["tolerance"] = out["tolerance"], ""
    return out


def _merge_llm_parse_with_ocr(llm_parsed: dict[str, str], ocr_base: dict[str, str]) -> dict[str, str]:
    """Keep LLM fields only when supported by OCR; never invent view/inspection metadata."""
    ocr_text = ocr_base.get("raw_ocr") or ocr_base.get("others") or ""
    out = dict(ocr_base)
    for key in ("nominal_value", "tolerance"):
        lv = (llm_parsed.get(key) or "").strip()
        if lv and _llm_value_supported_by_ocr(lv, ocr_text):
            out[key] = lv
    lo = (llm_parsed.get("others") or "").strip()
    if lo and _llm_value_supported_by_ocr(lo, ocr_text):
        out["others"] = lo[:2000]
    if os.environ.get("BALLOON_LLM_INCLUDE_META", "").strip().lower() in ("1", "true", "yes"):
        for key in ("feature_type", "view_location", "inspection_method", "remarks"):
            lv = (llm_parsed.get(key) or "").strip()
            if lv and _llm_value_supported_by_ocr(lv, ocr_text):
                out[key] = lv
    return out


def _enrich_extraction_from_ocr(
    parsed: dict[str, str],
    bgr: np.ndarray,
    class_name: str,
    orientation: str | None = None,
) -> dict[str, str]:
    """OCR-first; Claude vision OCR is authoritative when enabled."""
    ocr_base = _ocr_first_parse_crop(bgr, class_name, orientation)
    if _ocr_engine() == "claude":
        return ocr_base
    if not _vision_api_configured():
        return ocr_base
    has_llm = any((parsed.get(k) or "").strip() for k in ("nominal_value", "tolerance", "others"))
    if not has_llm:
        return ocr_base
    return _merge_llm_parse_with_ocr(parsed, ocr_base)


def _balloon_has_extracted_data(item: dict) -> bool:
    """Require at least one of: crop image, nominal value, or tolerance."""
    nom = str((item or {}).get("nominal_value") or "").strip()
    tol = str((item or {}).get("tolerance") or "").strip()
    crop = str(
        (item or {}).get("crop_preview_base64") or (item or {}).get("crop_save_base64") or ""
    ).strip()
    if len(crop) > 40:
        return True
    if nom and nom.lower() != "empty":
        return True
    if tol and tol.lower() != "empty":
        return True
    return False


def _hide_incomplete_balloons(payload: dict) -> None:
    """Remove balloons with no crop, nominal, or tolerance from UI and report."""
    _sync_balloon_items_from_detections(payload)
    items = list(payload.get("balloon_items") or [])
    skip_di: set[int] = set()
    kept: list = []
    for it in items:
        if _balloon_has_extracted_data(it):
            kept.append(it)
            continue
        di = _detection_index_for_item(payload, it)
        if di is not None:
            skip_di.add(int(di))
    payload["balloon_items"] = kept
    anns = list(payload.get("drawing_annotations") or [])
    for i in range(len(anns)):
        if i in skip_di:
            row = dict(anns[i])
            row["canvas_skip"] = True
            anns[i] = row
    payload["drawing_annotations"] = anns
    payload["balloons_hidden_no_data"] = len(skip_di)


def _detected_text_from_fields(nominal: str, tolerance: str) -> str:
    n = (nominal or "").strip()
    t = (tolerance or "").strip()
    if n and t:
        if t.startswith("±") or t.startswith("+") or t.startswith("-"):
            return f"{n} {t}"
        return f"{n} ± {t}"
    return n or t


def _title_block_meta_from_text(text: str) -> dict[str, str]:
    """Best-effort parse of title-block OCR text into part metadata fields."""
    blob = re.sub(r"\s+", " ", (text or "").strip())
    if not blob:
        return {}

    def pick(patterns: list[str]) -> str:
        for pat in patterns:
            m = re.search(pat, blob, re.IGNORECASE)
            if m and m.group(1).strip():
                return m.group(1).strip()
        return ""

    return {
        "part_number": pick(
            [
                r"(?:part\s*(?:no|number|#)|drawing\s*(?:no|number|#)|dwg\s*(?:no|#)?)\s*[:.\-]?\s*([A-Za-z0-9][A-Za-z0-9\-_/]*)",
                r"\b(P[-_]?\d[\w\-/]*)\b",
            ]
        ),
        "part_name": pick(
            [
                r"(?:part\s*name|description|title)\s*[:.\-]?\s*([A-Za-z0-9][\w\s\-/,]{2,80})",
            ]
        ),
        "revision": pick(
            [
                r"(?:rev(?:ision)?|issue)\s*[:.\-]?\s*([A-Za-z0-9]+)",
            ]
        ),
        "material": pick(
            [
                r"(?:material|matl)\s*[:.\-]?\s*([A-Za-z0-9][\w\s\-/,]{2,60})",
            ]
        ),
        "mass": pick(
            [
                r"(?:mass|weight)\s*[:.\-]?\s*([0-9][\d.,]*\s*(?:kg|g|lb|lbs)?)",
            ]
        ),
        "finish_treatment": pick(
            [
                r"(?:finish(?:\s*treatment)?|surface\s*finish|coating|treatment)\s*[:.\-]?\s*([A-Za-z0-9][\w\s\-/,]{2,80})",
            ]
        ),
    }


def _title_block_crop_from_image(img: np.ndarray) -> Optional[np.ndarray]:
    """Title block is usually bottom-right on engineering drawings."""
    if img is None or not getattr(img, "size", 0):
        return None
    h, w = img.shape[:2]
    x1 = int(w * 0.52)
    y1 = int(h * 0.62)
    x2 = w
    y2 = h
    if x2 - x1 < 80 or y2 - y1 < 60:
        return None
    return img[y1:y2, x1:x2].copy()


def _vision_title_block_json(crop: np.ndarray) -> dict[str, str]:
    """OpenAI / vision LLM extraction of the six inspection-report header fields."""
    parsed: dict[str, str] = {}
    ok, buf = cv2.imencode(".jpg", crop, [int(cv2.IMWRITE_JPEG_QUALITY), 92])
    if not ok:
        return parsed
    try:
        val = _tasks()._vision_llm_message(
            buf.tobytes(),
            (
                "You are reading an engineering drawing TITLE BLOCK (OCR-quality accuracy required).\n"
                "Extract these fields exactly as printed on the drawing:\n"
                "1. Part Number (drawing/part no., DWG NO)\n"
                "2. Part Name (title, description, name)\n"
                "3. Revision (REV, issue)\n"
                "4. Material (MATL)\n"
                "5. Mass or Weight (with units if shown)\n"
                "6. Finish Treatment (surface finish, coating, heat treat, etc.)\n"
                'Return ONLY valid JSON, no markdown:\n'
                '{"part_number":"","part_name":"","revision":"","material":"","mass":"","finish_treatment":""}\n'
                "Use empty string only when the field is truly not visible."
            ),
            max_tokens=800,
            temperature=0.0,
            top_p=1.0,
        )
        if val and not str(val).startswith("VISION_LLM_FAILED"):
            t = str(val).strip()
            if t.startswith("```"):
                t = re.sub(r"^```(?:json)?\s*", "", t, flags=re.IGNORECASE)
                t = re.sub(r"\s*```$", "", t)
            try:
                data = json.loads(t)
                if isinstance(data, dict):
                    for key in (
                        "part_number",
                        "part_name",
                        "revision",
                        "material",
                        "mass",
                        "finish_treatment",
                    ):
                        if data.get(key):
                            parsed[key] = str(data.get(key)).strip()
            except json.JSONDecodeError:
                pass
    except Exception:
        pass
    return parsed


def _merge_title_meta(base: dict[str, str], extra: dict[str, str]) -> dict[str, str]:
    out = dict(base)
    for key, val in (extra or {}).items():
        if val and not out.get(key):
            out[key] = val
    return out


def _extract_title_block_meta(image_path: str, detections: list) -> dict[str, str]:
    """Vision/OCR on YOLO Title_Block crop or bottom-right title block region."""
    img = _imread_bgr(image_path)
    if img is None:
        return {}
    h, w = img.shape[:2]
    crops: list[np.ndarray] = []

    title_det = None
    for d in detections or []:
        cls = (d.get("class_name") or "").strip().lower().replace(" ", "_")
        if "title" in cls and "block" in cls:
            title_det = d
            break
    if not title_det:
        for d in detections or []:
            cls = (d.get("class_name") or "").strip().lower()
            if cls in ("title_block", "titleblock", "title"):
                title_det = d
                break

    if title_det:
        bb = title_det.get("bbox") or []
        if len(bb) >= 4:
            x1, y1, x2, y2 = [int(v) for v in bb[:4]]
            x1 = max(0, min(w - 1, x1))
            y1 = max(0, min(h - 1, y1))
            x2 = max(x1 + 1, min(w, x2))
            y2 = max(y1 + 1, min(h, y2))
            crops.append(img[y1:y2, x1:x2])

    fallback = _title_block_crop_from_image(img)
    if fallback is not None:
        crops.append(fallback)

    parsed: dict[str, str] = {}
    for crop in crops:
        if not _deploy_safe_mode():
            parsed = _merge_title_meta(parsed, _vision_title_block_json(crop))
        ocr_text = _ocr_text_from_bgr(crop)
        if ocr_text:
            parsed = _merge_title_meta(parsed, _title_block_meta_from_text(ocr_text))
        if len([v for v in parsed.values() if v]) >= 4:
            break
    return parsed


def _bgr_from_jpeg_data_url_or_b64(s: str) -> Optional[np.ndarray]:
    raw = (s or "").strip()
    if not raw:
        return None
    if raw.startswith("data:"):
        i = raw.find(",")
        if i >= 0:
            raw = raw[i + 1 :]
    try:
        data = base64.standard_b64decode(raw)
    except Exception:
        return None
    arr = np.frombuffer(data, dtype=np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    return img


def _extract_one_crop_bgr_llm(bgr: np.ndarray, class_name: str) -> dict[str, str]:
    """Same JSON extraction as _extract_detection_text_llm for a single BGR crop."""
    nominal_value = ""
    tolerance = ""
    others = ""
    if bgr is None or not getattr(bgr, "size", 0):
        return {
            "nominal_value": nominal_value,
            "tolerance": tolerance,
            "others": others,
            "detected_text": "",
        }
    ch, cw = bgr.shape[:2]
    dim_ori = _dimension_callout_orientation(float(cw), float(ch))
    cls = (class_name or "").strip()
    if _ocr_engine() == "claude":
        parsed = _ocr_first_parse_crop(bgr, cls, dim_ori)
        nominal_value = parsed["nominal_value"]
        tolerance = parsed["tolerance"]
        others = parsed["others"]
        return {
            "nominal_value": nominal_value,
            "tolerance": tolerance,
            "others": others,
            "detected_text": _detected_text_from_fields(nominal_value, tolerance),
        }
    ok, buf = cv2.imencode(".jpg", _deskew_crop_for_vision(bgr, dim_ori), [int(cv2.IMWRITE_JPEG_QUALITY), 90])
    if ok:
        try:
            val = _tasks()._vision_llm_message(
                buf.tobytes(),
                _balloon_vision_prompt(cls),
                max_tokens=500,
                temperature=0.0,
                top_p=1.0,
            )
            parsed = _enrich_extraction_from_ocr(
                _parse_extraction_json(val or ""), bgr, cls, dim_ori
            )
            nominal_value = parsed["nominal_value"]
            tolerance = parsed["tolerance"]
            others = parsed["others"]
        except Exception:
            parsed = _enrich_extraction_from_ocr(
                {"nominal_value": "", "tolerance": "", "others": ""}, bgr, cls, dim_ori
            )
            nominal_value = parsed["nominal_value"]
            tolerance = parsed["tolerance"]
            others = parsed["others"]
    else:
        parsed = _enrich_extraction_from_ocr(
            {"nominal_value": "", "tolerance": "", "others": ""}, bgr, cls, dim_ori
        )
        nominal_value = parsed["nominal_value"]
        tolerance = parsed["tolerance"]
        others = parsed["others"]
    return {
        "nominal_value": nominal_value,
        "tolerance": tolerance,
        "others": others,
        "detected_text": _detected_text_from_fields(nominal_value, tolerance),
    }


def _crop_image_data_url(bgr, max_side: int = 320, min_side: int = 48) -> str:
    """
    Encode a BGR bbox crop as a data URL for the Others column (embedded in JSON as crop_preview_base64).
    Large crops are shrunk; very small YOLO boxes are upscaled so the thumbnail is visible (not an empty box).
    """
    if bgr is None or not getattr(bgr, "size", 0):
        return ""
    h, w = bgr.shape[:2]
    if h < 1 or w < 1:
        return ""
    # Upscale tiny crops so the UI shows a real image, not a blank sliver
    if min(h, w) < min_side:
        s = min_side / min(h, w)
        nw = max(1, int(round(w * s)))
        nh = max(1, int(round(h * s)))
        bgr = cv2.resize(bgr, (nw, nh), interpolation=cv2.INTER_LINEAR)
        h, w = bgr.shape[:2]
    if max(h, w) > max_side:
        s = max_side / max(h, w)
        nw = max(1, int(round(w * s)))
        nh = max(1, int(round(h * s)))
        bgr = cv2.resize(bgr, (nw, nh), interpolation=cv2.INTER_AREA)
    ok, buf = cv2.imencode(".jpg", bgr, [int(cv2.IMWRITE_JPEG_QUALITY), 88])
    if ok:
        return "data:image/jpeg;base64," + base64.b64encode(buf.tobytes()).decode("ascii")
    ok, buf = cv2.imencode(".png", bgr)
    if ok:
        return "data:image/png;base64," + base64.b64encode(buf.tobytes()).decode("ascii")
    return ""


def _exact_crop_jpeg_data_url(bgr, max_side: int = 8192) -> str:
    """
    Encode the bbox crop as JPEG at ~full resolution — no thumbnail downscale — for Save.
    Only scales down if a side exceeds max_side (safety for huge drawings).
    """
    if bgr is None or not getattr(bgr, "size", 0):
        return ""
    h, w = bgr.shape[:2]
    if h < 1 or w < 1:
        return ""
    if max(h, w) > max_side:
        s = max_side / max(h, w)
        nw = max(1, int(round(w * s)))
        nh = max(1, int(round(h * s)))
        bgr = cv2.resize(bgr, (nw, nh), interpolation=cv2.INTER_AREA)
    ok, buf = cv2.imencode(".jpg", bgr, [int(cv2.IMWRITE_JPEG_QUALITY), 95])
    if not ok:
        ok, buf = cv2.imencode(".png", bgr)
        if ok:
            return "data:image/png;base64," + base64.b64encode(buf.tobytes()).decode("ascii")
        return ""
    return "data:image/jpeg;base64," + base64.b64encode(buf.tobytes()).decode("ascii")


def _extract_detection_text_llm(image_path: str, detections: list) -> list:
    """
    Vision LLM on each YOLO bbox crop only (never the full sheet). Returns structured fields
    plus a JPEG data URL thumbnail of the crop for the UI.
    """
    img = _imread_bgr(image_path)
    if img is None:
        return []
    h, w = img.shape[:2]
    items = []
    max_ocr = _max_crop_ocr_count()
    if _deploy_safe_mode():
        print(f"[detect] Render safe mode: OCR on first {max_ocr} crops only (set BALLOON_MAX_CROP_OCR).")
    for i, d in enumerate(detections or [], start=1):
        bb = d.get("bbox") or []
        if len(bb) < 4:
            continue
        cls = (d.get("class_name") or "").strip()
        x1, y1, x2, y2 = [int(v) for v in bb[:4]]
        # Exact YOLO box (same as green rectangle in full image space) — no padding.
        x1 = max(0, min(w - 1, x1))
        y1 = max(0, min(h - 1, y1))
        x2 = max(x1 + 1, min(w, x2))
        y2 = max(y1 + 1, min(h, y2))
        if x2 <= x1 or y2 <= y1:
            continue
        crop = img[y1:y2, x1:x2]
        crop_h, crop_w = crop.shape[:2]
        dim_ori = (d.get("dimension_orientation") or "").strip() or _dimension_callout_orientation(
            float(crop_w), float(crop_h)
        )
        crop_preview_base64 = _crop_image_data_url(crop)
        crop_save_base64 = _exact_crop_jpeg_data_url(crop)
        parsed = {
            "nominal_value": "",
            "tolerance": "",
            "others": "",
            "feature_type": "",
            "view_location": "",
            "inspection_method": "",
            "remarks": "",
        }
        run_ocr = max_ocr <= 0 or i <= max_ocr
        if run_ocr:
            parsed = _ocr_first_parse_crop(crop, cls, dim_ori)
        nominal_value = parsed["nominal_value"]
        tolerance = parsed["tolerance"]
        others = parsed["others"]
        feature_type = parsed.get("feature_type") or ""
        view_location = parsed.get("view_location") or ""
        inspection_method = parsed.get("inspection_method") or ""
        remarks = parsed.get("remarks") or ""
        dim_text = _detected_text_from_fields(nominal_value, tolerance)
        detected_text = dim_text or (others or "").strip()
        region_name = (d.get("region_name") or "").strip()
        if not view_location and region_name:
            view_location = region_name
        items.append(
            {
                "balloon_number": i,
                "detection_index": i - 1,
                "class_name": cls,
                "confidence": d.get("confidence", ""),
                "feature_type": feature_type,
                "view_location": view_location,
                "inspection_method": inspection_method,
                "remarks": remarks,
                "nominal_value": nominal_value,
                "tolerance": tolerance,
                "others": others,
                "raw_ocr": (parsed.get("raw_ocr") or "").strip(),
                "ocr_engine": _ocr_engine(),
                "confirmed": False,
                "region_name": region_name,
                "detected_text": detected_text,
                "bbox_pixels": [x1, y1, x2, y2],
                "crop_preview_base64": crop_preview_base64,
                "crop_save_base64": crop_save_base64,
            }
        )
    return items


def _full_drawing_analysis_enabled() -> bool:
    if _deploy_safe_mode():
        return False
    return os.environ.get("BALLOON_FULL_DRAWING_ANALYSIS", "").strip().lower() in (
        "1",
        "true",
        "yes",
        "on",
    )


def _apply_full_drawing_balloon_analysis(payload: dict, image_path: str) -> None:
    """Optional Claude pass: full balloon table + CTQ / missing-dimension summaries."""
    if not _full_drawing_analysis_enabled() or not _vision_api_configured():
        return
    if not image_path or not Path(str(image_path)).is_file():
        return
    max_side = int(os.environ.get("BALLOON_VISION_MAX_SIDE", "2048"))
    try:
        image_bytes, _, _, _ = _prepare_image_bytes_for_vision(str(image_path), max_side)
    except Exception as exc:
        payload["full_drawing_analysis_error"] = str(exc)[:300]
        return
    title_meta = payload.get("title_block_meta") or {}
    tb = json.dumps(title_meta, ensure_ascii=False)[:4000] if title_meta else ""
    prompt = _mechanical_ballooning_prompts().full_drawing_analysis_prompt(tb)
    raw = _vision_llm_chat_direct(
        image_bytes,
        prompt,
        max_tokens=int(os.environ.get("BALLOON_FULL_ANALYSIS_MAX_TOKENS", "12000")),
        temperature=0.1,
    )
    if str(raw).strip().startswith("VISION_LLM_FAILED"):
        payload["full_drawing_analysis_error"] = str(raw).strip()[:500]
        return
    parsed = _parse_json_object_from_llm(raw)
    if isinstance(parsed, dict):
        payload["full_drawing_analysis"] = parsed
        payload["extraction_prompt_version"] = "mechanical_engineer_v1"


@app.get("/health")
async def health():
    return {"ok": True, "service": "serve_balloon", "port_hint": "default 10000"}


@app.get("/api/diagnostics")
async def api_diagnostics():
    """Quick checks: DB reachability and which UI folder is served."""
    return {
        "ok": True,
        "database_configured": bool(_Db),
        "mongodb_ping": db.ping(),
        "ui_path": str(_UI_DIR.resolve()),
    }


@app.get("/")
async def root_redirect():
    if balloon_auth_disabled():
        return RedirectResponse("/app")
    return RedirectResponse("/login")


@app.get("/login")
async def login_page():
    p = _UI_DIR / "login.html"
    if not p.is_file():
        raise HTTPException(500, "Missing login.html")
    return _html_no_cache(p)


@app.get("/payment")
async def payment_page():
    p = _UI_DIR / "payment.html"
    if not p.is_file():
        raise HTTPException(500, "Missing payment.html")
    return _html_no_cache(p)


@app.get("/change-password")
async def change_password_page():
    p = _UI_DIR / "change_password.html"
    if not p.is_file():
        raise HTTPException(500, "Missing change_password.html")
    return _html_no_cache(p)


@app.get("/app")
async def app_page():
    index = _UI_DIR / "index.html"
    if not index.is_file():
        return JSONResponse(status_code=500, content={"ok": False, "error": "Missing index.html"})
    return _html_no_cache(index)


@app.get("/admin")
async def admin_page():
    p = _UI_DIR / "admin.html"
    if not p.is_file():
        raise HTTPException(500, "Missing admin.html")
    return _html_no_cache(p)


@app.get("/inspection-report")
async def inspection_report_page():
    p = _UI_DIR / "inspection_report.html"
    if not p.is_file():
        raise HTTPException(500, "Missing inspection_report.html")
    return _html_no_cache(p)


@app.get("/api/v1/auth-config")
def api_auth_config():
    """Public config for static login/app pages."""
    amount_paise = int(os.environ.get("RAZORPAY_AMOUNT_PAISE", "99900"))
    db_ok = database_configured()
    dev_bypass = balloon_auth_disabled()
    return {
        "auth_enabled": auth_enabled(),
        "require_login": not dev_bypass,
        "dev_mode": dev_bypass,
        "database_configured": db_ok,
        "trial_days": trial_days(),
        "plan_amount_inr": amount_paise // 100,
        "payment_configured": bool(
            os.environ.get("RAZORPAY_KEY_ID", "").strip()
            and os.environ.get("RAZORPAY_KEY_SECRET", "").strip()
        ),
        "otp_email_configured": bool(os.environ.get("RESEND_API_KEY", "").strip()),
        "vision_fallback_available": _vision_api_configured(),
        "vision_fallback_provider": _vision_llm_provider(),
        "vision_fallback_mode": _vision_fallback_mode(),
    }


def _run_detection_pipeline(dest_path: str, work_dir: str, filename: str) -> tuple[dict | None, str | None]:
    """Heavy sync pipeline (YOLO + optional Claude). Runs in a worker thread on Render."""
    payload, err = _tasks().run_drawing_yolo_detection(dest_path, work_dir, filename)
    if err:
        return None, err

    payload["yolo_raw_count"] = payload.get("yolo_raw_count") or payload.get("count")
    _filter_detection_payload(payload)
    _mark_yolo_detection_sources(payload)
    payload["yolo_after_filter_count"] = payload.get("count")
    payload["pipeline_stage"] = "yolo_complete"
    _anthropic_region_prepass(payload)
    _opencv_dim_line_stage(payload)
    _apply_vision_fallback_if_needed(payload)
    _anthropic_coverage_verify(payload)
    _refine_dimension_detection_payload(payload)
    _reorder_detection_payload_tblr(payload)
    dets = payload.get("detections") or []
    payload["drawing_annotations"] = _drawing_annotations_from_detections(dets)
    extract_path = payload.get("infer_image_path") or dest_path
    dets_for_crop = payload.get("detections_full") or dets
    payload["balloon_items"] = _extract_detection_text_llm(extract_path, dets_for_crop)
    _expand_multiplier_balloons_payload(payload)
    _remove_duplicate_yolo_detections_near_multiplier(payload)
    _repair_multiplier_drawing_annotations(payload)
    if _balloon_placement_mode() in ("legacy", "autoballoon", "grid"):
        _apply_legacy_balloon_coordinates(payload, extract_path)
    else:
        payload["balloon_placement"] = "tight"
    _sync_balloon_items_from_detections(payload)
    _hide_incomplete_balloons(payload)
    payload["balloon_pipeline_complete"] = True
    payload["title_block_meta"] = _extract_title_block_meta(extract_path, dets_for_crop)
    payload["extraction_prompt"] = (
        "yolo_render_safe" if _deploy_safe_mode() else "yolo_balloons_then_claude_grid_gap_fill_ocr"
    )
    payload["render_safe_mode"] = _deploy_safe_mode()
    payload["show_detection_boxes"] = False
    if _full_drawing_analysis_enabled():
        _apply_full_drawing_balloon_analysis(payload, extract_path)
    payload["weights_path"] = _tasks().get_yolo_weights_path_loaded()
    return payload, None


@app.post("/api/v1/detect")
async def api_detect(
    file: UploadFile = File(...),
    current_user: User = Depends(require_balloon_write_access),
    pg: Session = Depends(get_db),
):
    if not file.filename:
        raise HTTPException(422, "No filename")

    suffix = Path(file.filename).suffix.lower()
    allowed = {".pdf", ".png", ".jpg", ".jpeg", ".bmp", ".webp", ".tif", ".tiff"}
    if suffix not in allowed:
        raise HTTPException(400, f"Unsupported type {suffix}. Use PDF or image.")

    job = uuid.uuid4().hex
    work = _UPLOAD_ROOT / job
    work.mkdir(parents=True, exist_ok=True)
    dest = work / f"input{suffix}"

    try:
        with dest.open("wb") as buf:
            shutil.copyfileobj(file.file, buf)

        try:
            payload, err = await asyncio.to_thread(
                _run_detection_pipeline, str(dest), str(work), file.filename
            )
        except Exception as exc:
            print(f"[detect] Pipeline failed: {exc}")
            return JSONResponse(
                status_code=500,
                content={"ok": False, "error": f"Detection failed: {exc}"[:500]},
            )

        if err:
            return JSONResponse(status_code=400, content={"ok": False, "error": err})

        dets = payload.get("detections") or []
        _log_activity(
            pg,
            current_user,
            action_type="drawing_upload",
            metadata={
                "filename": file.filename,
                "detection_count": len(dets),
                "balloon_count": len(payload.get("balloon_items") or []),
            },
        )

        return JSONResponse(
            content={
                "ok": True,
                "version": 1,
                "filename": file.filename,
                "detection": payload,
            }
        )
    finally:
        shutil.rmtree(work, ignore_errors=True)


@app.post("/api/v1/extract-balloon-text")
async def api_extract_balloon_text(
    body: ExtractBalloonTextBody,
    current_user: User = Depends(require_balloon_write_access),
    pg: Session = Depends(get_db),
):
    """
    Run vision LLM on a single crop (manual balloon box). Fills nominal_value / tolerance / others
    like automatic detection.
    """
    bgr = _bgr_from_jpeg_data_url_or_b64(body.crop_jpeg_base64)
    if bgr is None or not getattr(bgr, "size", 0):
        raise HTTPException(status_code=400, detail="Invalid or empty crop image")
    out = _extract_one_crop_bgr_llm(bgr, body.class_name)
    _log_activity(pg, current_user, action_type="balloon_text_extract", metadata={"class_name": body.class_name})
    return {"ok": True, "extract": out}


@app.post("/api/v1/export-excel")
async def api_export_excel(
    request: Request,
    current_user: User = Depends(require_balloon_write_access),
    pg: Session = Depends(get_db),
):
    payload = await request.json()
    detection = payload.get("detection") or {}
    filename = payload.get("filename") or "drawing"

    wb = Workbook()
    # ws_meta = wb.active
    # ws_meta.title = "summary"
    # ws_meta.append(["filename", filename])
    # ws_meta.append(["count", detection.get("count", 0)])
    # ws_meta.append(["width", detection.get("width", "")])
    # ws_meta.append(["height", detection.get("height", "")])
    # ws_meta.append(["input_kind", detection.get("input_kind", "")])
    # ws_meta.append(["weights_path", detection.get("weights_path", "")])

    # ws_det = wb.create_sheet("detections")
    # ws_det.append(["id", "class_name", "confidence", "x1", "y1", "x2", "y2"])
    # for idx, d in enumerate(detection.get("detections") or [], start=1):
    #     bb = d.get("bbox") or [None, None, None, None]
    #     ws_det.append(
    #         [
    #             idx,
    #             d.get("class_name", ""),
    #             d.get("confidence", ""),
    #             bb[0] if len(bb) > 0 else "",
    #             bb[1] if len(bb) > 1 else "",
    #             bb[2] if len(bb) > 2 else "",
    #             bb[3] if len(bb) > 3 else "",
    #         ]
    #     )

    # ws_ann = wb.create_sheet("balloons")
    # ws_ann.append(["id", "AnnotationType", "bbox_x1", "bbox_y1", "bbox_x2", "bbox_y2", "text_x", "text_y"])
    # for a in detection.get("drawing_annotations") or []:
    #     bb = a.get("BBox") or [None, None, None, None]
    #     tp = a.get("TextPos") or [None, None]
    #     ws_ann.append(
    #         [
    #             a.get("id", ""),
    #             a.get("AnnotationType", ""),
    #             bb[0] if len(bb) > 0 else "",
    #             bb[1] if len(bb) > 1 else "",
    #             bb[2] if len(bb) > 2 else "",
    #             bb[3] if len(bb) > 3 else "",
    #             tp[0] if len(tp) > 0 else "",
    #             tp[1] if len(tp) > 1 else "",
    #         ]
    #     )

    ws_items = wb.active
    ws_items.title = "balloon_items"
    ws_items.append(
        ["balloon_number", "class_name", "nominal_value", "tolerance", "others"]
    )
    for it in detection.get("balloon_items") or []:
        ws_items.append(
            [
                it.get("balloon_number", ""),
                it.get("class_name", ""),
                # it.get("confidence", ""),  # excluded from export
                it.get("nominal_value", ""),
                it.get("tolerance", ""),
                it.get("others", "") or it.get("detected_text", ""),
            ]
        )

    buff = BytesIO()
    wb.save(buff)
    xlsx_name = f"AutoBallooning_{Path(filename).stem}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{xlsx_name}"'}

    # ── Activity log ───────────────────────────────────────────────────────
    _log_activity(
        pg,
        current_user,
        action_type="excel_export",
        metadata={
            "filename": filename,
            "xlsx_name": xlsx_name,
            "row_count": len(detection.get("balloon_items") or []),
        },
    )

    return Response(
        content=buff.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.post("/api/v1/export-inspection-report-pdf")
@app.post("/api/export-inspection-report-pdf")
async def api_export_inspection_report_pdf(
    body: InspectionReportExportBody,
    current_user: User = Depends(require_balloon_write_access),
    pg: Session = Depends(get_db),
):
    """Build inspection report table + part meta as a downloadable PDF."""
    try:
        pdf_bytes = _build_inspection_report_pdf(body)
    except Exception as exc:
        return JSONResponse(
            status_code=500,
            content={"ok": False, "error": f"PDF generation failed: {exc}"},
        )

    stem = (body.part_number or "inspection_report").strip()
    stem = re.sub(r"[^\w\-]+", "_", stem) or "inspection_report"
    pdf_name = f"InspectionReport_{stem}.pdf"

    _log_activity(
        pg,
        current_user,
        action_type="inspection_report_pdf_export",
        metadata={"part_number": body.part_number, "row_count": len(body.rows)},
    )

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{pdf_name}"'},
    )


# ---------------------------------------------------------------------------
# Activity query endpoint
# ---------------------------------------------------------------------------
@app.get("/api/v1/activities")
def api_activities(
    tenant_id: Optional[str] = None,
    limit: int = 100,
    current_user: User = Depends(get_current_user),
    pg: Session = Depends(get_db),
):
    """
    Query the activity log.

    - Engineers see only their own tenant's activity.
    - Super admin sees all tenants (or filter by ?tenant_id=xxx).
    """
    from auth.schemas import ActivityResponse

    query = pg.query(Activity)

    if current_user.role == RoleEnum.super_admin:
        # Super admin: optionally filter by tenant
        if tenant_id:
            query = query.filter(Activity.tenant_id == tenant_id)
    else:
        # Engineer: always scoped to their own tenant
        query = query.filter(Activity.tenant_id == current_user.tenant_id)

    activities = query.order_by(Activity.created_at.desc()).limit(limit).all()
    return [ActivityResponse.model_validate(a) for a in activities]


# ---------------------------------------------------------------------------
# Trial status endpoint
# ---------------------------------------------------------------------------
@app.get("/api/v1/trial-status")
def api_trial_status(
    current_user: User = Depends(get_current_user),
    pg: Session = Depends(get_db),
):
    """
    Return the current tenant's subscription / trial status.
    Engineers call this to show the trial banner in the dashboard.
    Super admin always returns active.
    """
    if current_user.role == RoleEnum.super_admin:
        return {"subscription_status": "active", "is_active": True, "days_remaining": None}

    org = pg.query(Organization).filter_by(tenant_id=current_user.tenant_id).first()
    if not org or org.subscription_status is None:
        # Legacy tenant — treat as active
        return {"subscription_status": "active", "is_active": True, "days_remaining": None}

    # Potentially expire trial
    check_tenant_access(org, pg)

    days_remaining = None
    if org.subscription_status == "trial" and org.trial_end_date:
        trial_end = org.trial_end_date
        if trial_end.tzinfo is None:
            trial_end = trial_end.replace(tzinfo=timezone.utc)
        delta = trial_end - datetime.now(timezone.utc)
        days_remaining = max(0, delta.days)

    return {
        "subscription_status": org.subscription_status,
        "is_active": org.is_active if org.is_active is not None else True,
        "trial_start_date": org.trial_start_date.isoformat() if org.trial_start_date else None,
        "trial_end_date": org.trial_end_date.isoformat() if org.trial_end_date else None,
        "days_remaining": days_remaining,
    }


# ---------------------------------------------------------------------------
# Payment routes — Razorpay
# ---------------------------------------------------------------------------
@app.post("/payment/create-order")
def payment_create_order(
    current_user: User = Depends(get_current_user),
    pg: Session = Depends(get_db),
):
    """
    Create a Razorpay order for the current tenant.
    Requires RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET environment variables.
    """
    if current_user.role == RoleEnum.super_admin:
        raise HTTPException(status_code=400, detail="Super admin account does not require payment.")

    key_id = os.environ.get("RAZORPAY_KEY_ID", "").strip()
    key_secret = os.environ.get("RAZORPAY_KEY_SECRET", "").strip()
    amount_paise = int(os.environ.get("RAZORPAY_AMOUNT_PAISE", "99900"))  # default ₹999

    if not key_id or not key_secret:
        raise HTTPException(
            status_code=503,
            detail="Payment gateway is not configured. Contact the administrator.",
        )

    try:
        import razorpay  # noqa: F401 — installed via requirements.txt
        client = razorpay.Client(auth=(key_id, key_secret))
        order = client.order.create({
            "amount": amount_paise,
            "currency": "INR",
            "payment_capture": 1,
        })
        return {
            "order_id": order["id"],
            "amount": order["amount"],
            "currency": order["currency"],
            "key": key_id,
        }
    except ImportError:
        raise HTTPException(status_code=503, detail="Razorpay library not installed on server.")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to create payment order: {exc}")


@app.post("/payment/verify")
def payment_verify(
    body: PaymentVerifyRequest,
    current_user: User = Depends(get_current_user),
    pg: Session = Depends(get_db),
):
    """
    Verify Razorpay payment signature and activate the tenant's subscription.
    """
    if current_user.role == RoleEnum.super_admin:
        raise HTTPException(status_code=400, detail="Super admin account does not require payment.")

    key_secret = os.environ.get("RAZORPAY_KEY_SECRET", "").strip()
    if not key_secret:
        raise HTTPException(status_code=503, detail="Payment gateway is not configured.")

    # Verify HMAC-SHA256 signature as per Razorpay docs
    message = f"{body.razorpay_order_id}|{body.razorpay_payment_id}"
    expected_sig = hmac.new(
        key_secret.encode("utf-8"),
        message.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()

    if not hmac.compare_digest(expected_sig, body.razorpay_signature):
        raise HTTPException(status_code=400, detail="Invalid payment signature. Payment not verified.")

    # Activate subscription
    org = pg.query(Organization).filter_by(tenant_id=current_user.tenant_id).first()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found.")

    org.subscription_status = "active"
    org.is_active = True
    org.payment_id = body.razorpay_payment_id
    org.payment_date = datetime.now(timezone.utc)
    pg.commit()

    return {"status": "success", "message": "Subscription activated. Welcome to SmorX.ai!"}


if __name__ == "__main__":
    import uvicorn

    host = os.environ.get("BALLOON_UI_HOST", "127.0.0.1")
    port = int(os.environ.get("BALLOON_UI_PORT", "10000"))
    print(f"SmorX balloon UI + API  →  http://{host}:{port}/")
    print(f"App                     →  http://{host}:{port}/app")
    print(f"POST detection JSON     →  http://{host}:{port}/api/v1/detect")
    uvicorn.run(app, host=host, port=port, reload=False)
